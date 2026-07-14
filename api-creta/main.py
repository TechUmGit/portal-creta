"""
api-creta — Cloud Run service
Serve dados do BigQuery para o Portal Creta.

Endpoints:
  GET /api/receitas?periodo=12m   → todos os dados da página Receitas e Repasses
  GET /health                     → health check
"""

import io
import os
import re
import json
import time
import uuid
import base64
import logging
import requests
from datetime import datetime
from typing import Optional
from urllib.parse import quote
import pandas as pd
import yfinance as yf

from fastapi import FastAPI, Depends, HTTPException, Header, Query, UploadFile, File, Form, Request, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from google.cloud import bigquery, storage as gcs
from google.cloud.bigquery import ScalarQueryParameter, QueryJobConfig
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth, firestore as fb_firestore

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("api-creta")

# ── Firebase Admin (verifica tokens JWT do frontend) ─────────────────────────
# Em Cloud Run: a variável GOOGLE_APPLICATION_CREDENTIALS aponta para o arquivo
# de service account. Em dev local, use: firebase_admin.initialize_app() que
# detecta o ADC (Application Default Credentials).
if not firebase_admin._apps:
    firebase_admin.initialize_app(options={"projectId": "creta-btg-bd3a8"})

# ── BigQuery client ───────────────────────────────────────────────────────────
# Project ID e dataset — configuráveis via variável de ambiente
GCP_PROJECT    = os.getenv("GCP_PROJECT", "creta-btg")
bq = bigquery.Client(project=GCP_PROJECT)
DATASET        = os.getenv("BQ_DATASET",  "dados_crus")
TABLE          = f"`{GCP_PROJECT}.{DATASET}.receitas_para_repasse`"
TABLE_POSICAO     = f"`{GCP_PROJECT}.{DATASET}.posicao_das_contas`"
TABLE_SUITABILITY = f"`{GCP_PROJECT}.{DATASET}.suitability_contas`"
TABLE_EXCECOES      = f"`{GCP_PROJECT}.{DATASET}.conta_assessor_excecoes`"
TABLE_ASSESSOR_BASE = f"`{GCP_PROJECT}.{DATASET}.conta_assessor_base`"
TABLE_WEBHOOK_RAW   = f"{GCP_PROJECT}.{DATASET}.webhook_btg_raw"
TABLE_CR_ALLOCATION      = f"`{GCP_PROJECT}.{DATASET}.carteira_recomendada_allocation`"
TABLE_CR_PORTFOLIO       = f"`{GCP_PROJECT}.{DATASET}.carteira_recomendada_portfolio`"
TABLE_PRIMEIRA_APARICAO  = f"`{GCP_PROJECT}.{DATASET}.conta_primeira_aparicao`"
TABLE_CDB_LCA            = f"`{GCP_PROJECT}.{DATASET}.partner_report_cdb_lca_lci_lf`"

# ── GCS ───────────────────────────────────────────────────────────────────────
GCS_BUCKET         = os.getenv("GCS_BUCKET", "creta-btg-pipeline")

# ── NewsAPI (para refresh manual de notícias) ─────────────────────────────────
NEWSAPI_KEY             = os.getenv("NEWSAPI_KEY")          # secret montado no Cloud Run
NEWSAPI_URL_TUDO        = "https://newsapi.org/v2/everything"
NOTICIAS_LIMITE_MANUAL  = int(os.getenv("NOTICIAS_LIMITE_MANUAL", "5"))  # refreshes/dia
NOTICIAS_CATEGORIAS = [
    {"nome": "Renda Variável",  "slug": "rv",            "query": 'Ibovespa OR "bolsa de valores" OR B3'},
    {"nome": "Renda Fixa",      "slug": "rf",            "query": '"renda fixa" OR "tesouro direto" OR CDB OR LCI'},
    {"nome": "Macro",           "slug": "macro",         "query": 'PIB OR "banco central" OR "economia brasileira"'},
    {"nome": "Curva de Juros",  "slug": "juros",         "query": '"curva de juros" OR "juros futuros" OR "DI futuro" OR Selic'},
    {"nome": "IPCA",            "slug": "ipca",          "query": 'IPCA OR inflação OR INPC'},
    {"nome": "Internacional",   "slug": "internacional", "query": '"mercados internacionais" OR "wall street" OR fed OR dólar'},
]
# ── BTG OAuth2 (movimentações) ────────────────────────────────────────────────
BTG_CLIENT_ID     = os.getenv("BTG_CLIENT_ID")
BTG_CLIENT_SECRET = os.getenv("BTG_CLIENT_SECRET")
BTG_TOKEN_URL     = "https://api.btgpactual.com/iaas-auth/api/v1/authorization/oauth2/accesstoken"
BTG_MOV_URL       = "https://api.btgpactual.com/iaas-api-operation/api/v1/operation-history/full"
_btg_token_cache: dict = {"token": None, "expires_at": 0.0}

def _get_btg_token() -> str:
    agora = time.time()
    if _btg_token_cache["token"] and agora < _btg_token_cache["expires_at"] - 30:
        return _btg_token_cache["token"]
    if not BTG_CLIENT_ID or not BTG_CLIENT_SECRET:
        raise RuntimeError("BTG_CLIENT_ID / BTG_CLIENT_SECRET não configurados.")
    creds_b64 = base64.b64encode(f"{BTG_CLIENT_ID}:{BTG_CLIENT_SECRET}".encode()).decode()
    resp = requests.post(
        BTG_TOKEN_URL,
        headers={
            "Authorization":        f"Basic {creds_b64}",
            "Content-Type":         "application/x-www-form-urlencoded",
            "x-id-partner-request": str(uuid.uuid4()),
        },
        data={"grant_type": "client_credentials"},
        timeout=15,
    )
    resp.raise_for_status()
    # Token vem no header da resposta (padrão BTG)
    token = resp.headers.get("access_token")
    if not token:
        # fallback: tenta no body JSON
        try:
            token = resp.json().get("access_token")
        except Exception:
            pass
    if not token:
        raise RuntimeError(f"Token BTG não encontrado. Headers: {dict(resp.headers)}")
    _btg_token_cache["token"]      = token
    _btg_token_cache["expires_at"] = agora + 3600
    return token

GCS_PREFIX         = "entradas/"
GCS_PREFIX_PRODUTOS  = "produtos-manuais/"
GCS_PREFIX_CHAMADOS  = "chamados/"
gcs_client         = gcs.Client(project=GCP_PROJECT)

# ── Cotações via Yahoo Finance ────────────────────────────────────────────────
# Tickers BTG internos → código B3 padrão (.SA é o sufixo do Yahoo para B3)
TICKER_MAP_YF: dict[str, str] = {
    "VALEON": "VALE3.SA",
    "PETRPN": "PETR4.SA",
    "PETRPP": "PETR3.SA",
    "BRADPN": "BBDC4.SA",
    "RENTON": "RENT3.SA",
}
_preco_cache: dict = {}   # { ref: {"v": float|None, "ts": float} }
PRECO_TTL = 900           # 15 minutos

def _yf_ticker(ref: str) -> str:
    return TICKER_MAP_YF.get(ref, ref + ".SA")

def _ref_from_info(info: str | None) -> str | None:
    if not info:
        return None
    m = re.search(r"Ref:\s*([^\s|]+)", info)
    return m.group(1).strip() if m else None

def _yf_fetch_one(yft: str) -> float | None:
    """Busca preço de um único ticker via yfinance."""
    try:
        raw = yf.download(yft, period="5d", progress=False, auto_adjust=True)
        if raw.empty:
            return None
        col = raw["Close"].dropna()
        return round(float(col.iloc[-1]), 2) if not col.empty else None
    except Exception:
        return None

def buscar_precos(refs: list[str]) -> dict[str, float]:
    """Retorna {ref: preco} usando cache de 15 min. Busca via yfinance.
    Falhas NÃO são cacheadas — serão retentadas na próxima chamada.
    Se o batch falhar, tenta cada ticker individualmente.
    """
    agora = time.time()
    faltando = [r for r in set(refs)
                if r not in _preco_cache or (agora - _preco_cache[r]["ts"]) > PRECO_TTL]

    if faltando:
        tmap = {r: _yf_ticker(r) for r in faltando}
        uniq = list(set(tmap.values()))
        resultados: dict[str, float | None] = {}

        # 1ª tentativa: batch
        try:
            raw = yf.download(uniq, period="5d", progress=False, auto_adjust=True)
            if not raw.empty:
                if len(uniq) == 1:
                    col = raw["Close"].dropna()
                    v   = round(float(col.iloc[-1]), 2) if not col.empty else None
                    resultados[uniq[0]] = v
                else:
                    close = raw["Close"]
                    for yft in uniq:
                        try:
                            col = close[yft].dropna()
                            resultados[yft] = round(float(col.iloc[-1]), 2) if not col.empty else None
                        except Exception:
                            resultados[yft] = None
        except Exception as e:
            log.warning(f"yfinance batch erro: {e}")

        # 2ª tentativa: individual para os que falharam no batch
        falhou_batch = [yft for yft in uniq if resultados.get(yft) is None]
        if falhou_batch:
            log.info(f"yfinance retry individual: {falhou_batch}")
            for yft in falhou_batch:
                resultados[yft] = _yf_fetch_one(yft)

        # Cacheia apenas sucessos; falhas serão retentadas na próxima chamada
        for ref, yft in tmap.items():
            v = resultados.get(yft)
            if v is not None:
                _preco_cache[ref] = {"v": v, "ts": agora}
            else:
                log.warning(f"Preço não obtido para {ref} ({yft}) — não cacheando falha")

    return {r: _preco_cache[r]["v"] for r in set(refs)
            if _preco_cache.get(r, {}).get("v") is not None}


# ── Cache simples em memória ──────────────────────────────────────────────────
# Evita consultar o BigQuery a cada requisição.
# TTL padrão: 60 minutos (os dados chegam via webhook 1x por dia).
_cache: dict = {}
CACHE_TTL = int(os.getenv("CACHE_TTL_SECONDS", 3600))  # 1 hora

def cache_get(key: str):
    entry = _cache.get(key)
    if entry and (time.time() - entry["ts"]) < CACHE_TTL:
        log.info(f"Cache HIT: {key}")
        return entry["data"]
    return None

def cache_set(key: str, data):
    _cache[key] = {"data": data, "ts": time.time()}
    log.info(f"Cache SET: {key} ({len(str(data))} bytes)")

# ── FastAPI app ───────────────────────────────────────────────────────────────
app = FastAPI(title="API Creta Capital", version="1.0.0")

# CORS — permite chamadas do portal (localhost em dev, GitHub Pages em prod)
ALLOWED_ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:8080,http://127.0.0.1:8080"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

# ── Emails de admin (fallback caso Firestore/claims não estejam configurados) ──
ADMIN_EMAILS = {"fillipemsousa@gmail.com", "manu.lombardi@cretainvestimentos.com.br"}

# ── Firestore client (lazy, para buscar perfil quando claims não existem) ──────
_fs = None
def get_fs():
    global _fs
    if _fs is None:
        _fs = fb_firestore.client()
    return _fs

# ── Auth: verifica Bearer token do Firebase e retorna (role, assessor_name) ───
async def verificar_token(authorization: Optional[str] = None) -> dict:
    """
    Extrai e valida o Firebase ID token do header Authorization.
    Retorna dict com uid, email, role, assessor_name.
    """
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de autenticação ausente ou inválido."
        )
    token = authorization.split(" ", 1)[1]
    try:
        decoded = firebase_auth.verify_id_token(token)
    except Exception as e:
        log.warning(f"Token inválido: {e}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token expirado ou inválido. Faça login novamente."
        )

    uid   = decoded.get("uid", "")
    email = decoded.get("email", "")

    # 1ª opção: usar custom claims (definidos em criar_usuarios.py)
    role          = decoded.get("role")
    assessor_name = decoded.get("assessor_name")

    # 2ª opção: buscar no Firestore (usuários que não passaram pelo script)
    if not role:
        try:
            doc = get_fs().collection("users").document(uid).get()
            if doc.exists:
                perfil = doc.to_dict()
                role          = perfil.get("role")
                assessor_name = perfil.get("assessor_name")
        except Exception as fe:
            log.warning(f"Não foi possível ler perfil Firestore: {fe}")

    # 3ª opção: fallback por e-mail (contas antigas / dev)
    if not role:
        role = "admin" if email in ADMIN_EMAILS else "assessor"

    decoded["role"]          = role
    decoded["assessor_name"] = assessor_name
    return decoded

# ── Helpers SQL ───────────────────────────────────────────────────────────────
def where_periodo(periodo: str, start_date: str = None, end_date: str = None) -> str:
    """Gera cláusula WHERE para o filtro de período."""
    if periodo == "custom" and start_date:
        fim = f"DATE '{end_date}'" if end_date else "CURRENT_DATE()"
        return f"DATE(Data_De_Referencia) >= DATE '{start_date}' AND DATE(Data_De_Referencia) <= {fim}"
    mapa = {
        "3m":  "DATE_SUB(CURRENT_DATE(), INTERVAL 3  MONTH)",
        "6m":  "DATE_SUB(CURRENT_DATE(), INTERVAL 6  MONTH)",
        "12m": "DATE_SUB(CURRENT_DATE(), INTERVAL 12 MONTH)",
        "ytd": "DATE_TRUNC(CURRENT_DATE(), YEAR)",
        "all": "DATE '2000-01-01'",
    }
    data_inicio = mapa.get(periodo, mapa["12m"])
    return f"DATE(Data_De_Referencia) >= {data_inicio}"

def run_query(sql: str) -> list[dict]:
    """Executa uma query no BigQuery e retorna lista de dicts."""
    log.info(f"BQ query: {sql[:120]}...")
    rows = bq.query(sql).result()
    return [dict(row) for row in rows]

def fmt_brl(v: float, decimais: int = 2) -> str:
    """Formata valor no padrão brasileiro: 1.234.567,89"""
    s = f"{abs(v):,.{decimais}f}"          # "1,234,567.89" (en)
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')  # "1.234.567,89"
    return s

def _mapa_assessor_sq() -> str:
    """
    Subquery SQL que devolve (Conta INT64, Assessor STRING) para todas as contas.
    Prioridade: exceções ativas (conta_assessor_excecoes) >
                base mensal mais recente (conta_assessor_base).
    Pode ser usada inline como subquery em qualquer JOIN ou WHERE.
    """
    return f"""(
        SELECT Conta, Assessor
        FROM (
            SELECT e.Conta, e.Assessor, 1 AS prioridade
            FROM {TABLE_EXCECOES} e
            WHERE e.DataInicio <= CURRENT_DATE()
              AND (e.DataFim IS NULL OR e.DataFim >= CURRENT_DATE())
            UNION ALL
            SELECT b.Conta, b.Assessor, 2 AS prioridade
            FROM {TABLE_ASSESSOR_BASE} b
            WHERE b.MesRef = (SELECT MAX(MesRef) FROM {TABLE_ASSESSOR_BASE})
        )
        QUALIFY ROW_NUMBER() OVER (PARTITION BY Conta ORDER BY prioridade) = 1
    )"""

# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/receitas")
async def receitas(
    periodo: str = "12m",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna todos os dados da página Receitas e Repasses em uma única chamada.
    O frontend recebe um único JSON e monta todos os gráficos e tabelas.

    Parâmetro:
      periodo: "3m" | "6m" | "12m" | "ytd" | "all"  (default: "12m")
    """
    token_data = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")

    # Admins veem tudo; assessores veem apenas seus próprios dados
    is_admin        = role == "admin"
    forced_assessor = None if is_admin else assessor_name

    cache_key = f"receitas:{periodo}:{start_date or ''}:{end_date or ''}:{forced_assessor or 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    where = where_periodo(periodo, start_date, end_date)

    # Filtro extra para não-admins
    qp: list = []
    if forced_assessor:
        where += " AND UPPER(TRIM(Assessor_Manual)) = UPPER(@assessor)"
        qp.append(ScalarQueryParameter("assessor", "STRING", forced_assessor.strip()))

    def run_q(sql: str) -> list[dict]:
        if qp:
            rows = bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result()
        else:
            rows = bq.query(sql).result()
        log.info(f"BQ query: {sql[:100]}...")
        return [dict(r) for r in rows]

    # 1. KPIs — totais do período
    kpis = run_q(f"""
        SELECT
          ROUND(SUM(Comissao),             2) AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),     2) AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido),2) AS repasse_liquido,
          ROUND(SUM(Comissao_Liquida) - SUM(Repasse_Total_liquido), 2) AS receita_escritorio
        FROM {TABLE}
        WHERE {where}
    """)[0]

    # 2. Evolução mensal — barra principal
    evolucao = run_q(f"""
        SELECT
          FORMAT_DATE('%b/%y', DATE(Data_De_Referencia))      AS mes,
          DATE_TRUNC(DATE(Data_De_Referencia), MONTH)         AS data_ref,
          ROUND(SUM(Comissao),              2)                 AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2)                 AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2)                 AS repasse_liquido,
          ROUND(SUM(Comissao_Liquida) - SUM(Repasse_Total_liquido), 2) AS receita_escritorio
        FROM {TABLE}
        WHERE {where}
        GROUP BY mes, data_ref
        ORDER BY data_ref ASC
    """)
    # Converte data_ref para string para serialização JSON
    for row in evolucao:
        if hasattr(row.get("data_ref"), "isoformat"):
            row["data_ref"] = row["data_ref"].isoformat()

    # 3. Por assessor — donuts + tabela
    por_assessor = run_q(f"""
        SELECT
          UPPER(COALESCE(Assessor_Manual, 'Sem assessor')) AS assessor,
          ROUND(SUM(Comissao),              2)       AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2)       AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2)       AS repasse_liquido
        FROM {TABLE}
        WHERE {where}
          AND Assessor_Manual IS NOT NULL
        GROUP BY assessor
        ORDER BY repasse_liquido DESC
    """)

    # 4. Por categoria — tabela
    por_categoria = run_q(f"""
        SELECT
          COALESCE(Categoria_de_Repasse, 'Outros') AS categoria,
          ROUND(SUM(Comissao),               2)    AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),       2)    AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido),  2)    AS repasse
        FROM {TABLE}
        WHERE {where}
        GROUP BY categoria
        ORDER BY repasse DESC
    """)

    # 5. Por cliente — tabela (top 100 por repasse)
    por_cliente = run_q(f"""
        SELECT
          UPPER(COALESCE(Assessor_Manual, '')) AS assessor,
          COALESCE(Cliente, '')         AS cliente,
          ROUND(SUM(Receita_Bruta),              2) AS receita_bruta,
          ROUND(SUM(Receita_Liquida),            2) AS receita_liquida,
          ROUND(SUM(Comissao),                   2) AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),           2) AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido),      2) AS repasse
        FROM {TABLE}
        WHERE {where}
          AND Cliente IS NOT NULL
          AND Cliente != ''
        GROUP BY assessor, cliente
        ORDER BY repasse DESC
        LIMIT 100
    """)

    # 6. Evolução mensal por assessor (usada quando assessor/cliente está filtrado no frontend)
    evolucao_por_assessor = run_q(f"""
        SELECT
          UPPER(COALESCE(Assessor_Manual, '')) AS assessor,
          FORMAT_DATE('%b/%y', DATE(Data_De_Referencia))      AS mes,
          DATE_TRUNC(DATE(Data_De_Referencia), MONTH)         AS data_ref,
          ROUND(SUM(Comissao),              2)                 AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2)                 AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2)                 AS repasse_liquido,
          ROUND(SUM(Comissao_Liquida) - SUM(Repasse_Total_liquido), 2) AS receita_escritorio
        FROM {TABLE}
        WHERE {where}
          AND Assessor_Manual IS NOT NULL
        GROUP BY assessor, mes, data_ref
        ORDER BY assessor, data_ref ASC
    """)
    for row in evolucao_por_assessor:
        if hasattr(row.get("data_ref"), "isoformat"):
            row["data_ref"] = row["data_ref"].isoformat()

    # 7. Por categoria por assessor (usada quando assessor/cliente está filtrado no frontend)
    por_assessor_categoria = run_q(f"""
        SELECT
          UPPER(COALESCE(Assessor_Manual, '')) AS assessor,
          COALESCE(Categoria_de_Repasse, 'Outros') AS categoria,
          ROUND(SUM(Comissao),              2) AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2) AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2) AS repasse
        FROM {TABLE}
        WHERE {where}
          AND Assessor_Manual IS NOT NULL
        GROUP BY assessor, categoria
        ORDER BY assessor, repasse DESC
    """)

    resultado = {
        "periodo":               periodo,
        "gerado_em":             datetime.utcnow().isoformat(),
        "kpis":                  kpis,
        "evolucao":              evolucao,
        "por_assessor":          por_assessor,
        "por_categoria":         por_categoria,
        "por_cliente":           por_cliente,
        "evolucao_por_assessor": evolucao_por_assessor,
        "por_assessor_categoria": por_assessor_categoria,
    }

    cache_set(cache_key, resultado)
    return resultado


@app.get("/api/detalhe")
async def detalhe(
    periodo: str = "12m",
    assessor: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna dados linha a linha para o relatório de detalhe Excel.
    Parâmetros:
      periodo: "3m" | "6m" | "12m" | "ytd" | "all" | "custom"
      start_date / end_date: obrigatórios quando periodo="custom" (formato YYYY-MM-DD)
      assessor: nome exato do assessor (apenas admins podem especificar um diferente do seu)
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    # Não-admins só podem ver seus próprios dados
    if not is_admin:
        assessor = assessor_name  # ignora parâmetro da URL

    cache_key = f"detalhe:{periodo}:{start_date}:{end_date}:{assessor or 'todos'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    where = where_periodo(periodo, start_date, end_date)
    query_params = []

    if assessor:
        where += " AND UPPER(TRIM(Assessor_Manual)) = UPPER(@assessor)"
        query_params.append(ScalarQueryParameter("assessor", "STRING", assessor.strip()))

    sql = f"""
        SELECT
          FORMAT_DATE('%Y-%m-%d', DATE(Data_Receita))      AS Data_Receita,
          CAST(Conta AS STRING)                              AS Conta,
          COALESCE(Cliente, '')                              AS Cliente,
          COALESCE(Categoria_de_Repasse, '')                AS Categoria,
          COALESCE(Produto, '')                              AS Produto,
          COALESCE(Ativo, '')                                AS Ativo,
          COALESCE(Tipo_Receita, '')                         AS Tipo_Receita,
          ROUND(COALESCE(Receita_Bruta, 0), 4)              AS Receita_Bruta,
          ROUND(COALESCE(Receita_Liquida, 0), 4)            AS Receita_Liquida,
          ROUND(COALESCE(Comissao, 0), 4)                   AS Comissao,
          ROUND(COALESCE(Taxa_de_Repasse, 0), 4)            AS Taxa_de_Repasse,
          ROUND(COALESCE(Comissao_Liquida, 0), 4)           AS Comissao_Liquida,
          ROUND(COALESCE(Repasse_Total_liquido, 0), 4)      AS Repasse_Total_liquido
        FROM {TABLE}
        WHERE {where}
        ORDER BY Data_De_Referencia DESC, Cliente
        LIMIT 10000
    """

    job_config = QueryJobConfig(query_parameters=query_params) if query_params else None
    log.info(f"BQ detalhe query: periodo={periodo}, assessor={assessor}")
    if job_config:
        rows = bq.query(sql, job_config=job_config).result()
    else:
        rows = bq.query(sql).result()

    resultado = {"rows": [dict(r) for r in rows]}
    cache_set(cache_key, resultado)
    return resultado


@app.get("/api/evolucao_cliente")
async def evolucao_cliente(
    periodo: str = "12m",
    cliente: str = "",
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna evolução mensal de um cliente específico.
    Chamado pelo frontend quando o usuário seleciona um cliente no filtro.
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    if not cliente:
        raise HTTPException(status_code=400, detail="Parâmetro 'cliente' é obrigatório.")

    where  = where_periodo(periodo)
    qp     = [ScalarQueryParameter("cliente", "STRING", cliente.strip())]
    where += " AND TRIM(Cliente) = @cliente"

    # Não-admins só podem consultar clientes do próprio assessor
    if not is_admin and assessor_name:
        where += " AND UPPER(TRIM(Assessor_Manual)) = UPPER(@assessor)"
        qp.append(ScalarQueryParameter("assessor", "STRING", assessor_name.strip()))

    sql = f"""
        SELECT
          FORMAT_DATE('%b/%y', DATE(Data_De_Referencia))      AS mes,
          DATE_TRUNC(DATE(Data_De_Referencia), MONTH)         AS data_ref,
          ROUND(SUM(Comissao),              2)                 AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2)                 AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2)                 AS repasse_liquido,
          ROUND(SUM(Comissao_Liquida) - SUM(Repasse_Total_liquido), 2) AS receita_escritorio
        FROM {TABLE}
        WHERE {where}
        GROUP BY mes, data_ref
        ORDER BY data_ref ASC
    """
    sql_cat = f"""
        SELECT
          COALESCE(Categoria_de_Repasse, 'Outros') AS categoria,
          ROUND(SUM(Comissao),              2) AS comissao_bruta,
          ROUND(SUM(Comissao_Liquida),      2) AS comissao_liquida,
          ROUND(SUM(Repasse_Total_liquido), 2) AS repasse
        FROM {TABLE}
        WHERE {where}
        GROUP BY categoria
        ORDER BY repasse DESC
    """
    log.info(f"BQ evolucao_cliente: cliente={cliente!r}, periodo={periodo}")
    rows     = bq.query(sql,     job_config=QueryJobConfig(query_parameters=qp)).result()
    rows_cat = bq.query(sql_cat, job_config=QueryJobConfig(query_parameters=qp)).result()

    result = []
    for r in rows:
        row = dict(r)
        if hasattr(row.get("data_ref"), "isoformat"):
            row["data_ref"] = row["data_ref"].isoformat()
        result.append(row)

    return {"evolucao": result, "categorias": [dict(r) for r in rows_cat]}


@app.get("/api/posicoes")
async def posicoes_endpoint(
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna posições consolidadas por conta e classe, unindo:
      - posicao_das_contas   (data mais recente)
      - receitas_para_repasse (nome do cliente + assessor)
      - suitability_contas   (perfil de suitability mais recente)

    Admins veem todas as contas; assessores veem apenas suas próprias.
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"
    forced_assessor = None if is_admin else assessor_name

    cache_key = f"posicoes:{forced_assessor or 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    qp = []
    if forced_assessor:
        qp.append(ScalarQueryParameter("assessor", "STRING", forced_assessor.strip()))
        ma_join = f"INNER JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ma.Conta AND UPPER(ma.Assessor) = UPPER(@assessor)"
    else:
        ma_join = f"LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ma.Conta"

    sql = f"""
        WITH ultima_data AS (
            SELECT MAX(DATE(Data)) AS max_data
            FROM {TABLE_POSICAO}
        ),
        posicao_base AS (
            SELECT
                TRIM(p.Conta) AS Conta,
                p.Classe,
                ROUND(SUM(p.ValorBruto), 2) AS auc
            FROM {TABLE_POSICAO} p
            JOIN ultima_data ON DATE(p.Data) = ultima_data.max_data
            WHERE p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta), p.Classe
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE}
            WHERE Cliente IS NOT NULL AND Cliente != ''
            GROUP BY Conta
        ),
        suit AS (
            SELECT Conta, Perfil
            FROM {TABLE_SUITABILITY}
            QUALIFY ROW_NUMBER() OVER (PARTITION BY Conta ORDER BY DataExtracao DESC) = 1
        )
        SELECT
            p.Conta,
            COALESCE(n.cliente,  '')            AS cliente,
            UPPER(COALESCE(ma.Assessor, ''))           AS assessor,
            COALESCE(suit.Perfil, 'Sem perfil') AS perfil,
            p.Classe,
            p.auc,
            ud.max_data AS data_referencia
        FROM posicao_base p
        CROSS JOIN ultima_data ud
        {ma_join}
        LEFT JOIN nomes n  ON SAFE_CAST(TRIM(p.Conta) AS INT64) = n.conta_num
        LEFT JOIN suit     ON p.Conta = suit.Conta
        ORDER BY COALESCE(n.cliente, p.Conta), p.Classe
    """

    log.info(f"BQ posicoes: assessor={forced_assessor or 'todos'}")
    if qp:
        rows_raw = list(bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result())
    else:
        rows_raw = list(bq.query(sql).result())

    rows_list: list[dict] = []
    data_ref: Optional[str] = None
    por_classe: dict = {}

    for row in rows_raw:
        d = dict(row)
        dr = d.pop("data_referencia", None)
        if data_ref is None and dr is not None:
            data_ref = dr.isoformat() if hasattr(dr, "isoformat") else str(dr)
        classe = d.get("Classe") or "Outros"
        por_classe[classe] = round(por_classe.get(classe, 0) + (d.get("auc") or 0), 2)
        rows_list.append(d)

    por_classe_list = [
        {"classe": k, "auc": v}
        for k, v in sorted(por_classe.items(), key=lambda x: -x[1])
    ]

    resultado = {
        "data_referencia":   data_ref,
        "por_conta_classe":  rows_list,
        "por_classe":        por_classe_list,
    }
    cache_set(cache_key, resultado)
    return resultado


@app.get("/api/opcoes")
async def opcoes_endpoint(
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna todas as posições de derivativos (opções) da data mais recente.
    Admins veem todas as contas; assessores veem apenas as suas.
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    cache_key = f"opcoes:{assessor_name if not is_admin else 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    qp = []
    if not is_admin and assessor_name:
        qp.append(ScalarQueryParameter("assessor", "STRING", assessor_name.strip()))
        join_part = f"""
            INNER JOIN (
                SELECT CAST(Conta AS STRING) AS cnt
                FROM {_mapa_assessor_sq()}
                WHERE UPPER(Assessor) = UPPER(@assessor)
            ) ma ON TRIM(p.Conta) = ma.cnt
        """
    else:
        join_part = ""

    sql = f"""
        WITH ultima_data AS (
            SELECT MAX(DATE(Data)) AS max_data
            FROM {TABLE_POSICAO}
            WHERE Classe = 'Derivativo'
        ),
        clientes AS (
            SELECT
                Conta            AS conta_num,
                MAX(Cliente)     AS cliente
            FROM {TABLE}
            WHERE Cliente IS NOT NULL AND Cliente != ''
            GROUP BY Conta
        )
        SELECT
            p.Conta,
            COALESCE(c.cliente, '')  AS Cliente,
            p.Subclasse,
            p.Nome,
            p.Ticker,
            FORMAT_DATE('%Y-%m-%d', SAFE_CAST(p.Vencimento AS DATE)) AS Vencimento,
            p.Direcao,
            SAFE_CAST(p.Quantidade AS FLOAT64) AS Quantidade,
            SAFE_CAST(p.Preco      AS FLOAT64) AS Preco,
            SAFE_CAST(p.ValorBruto AS FLOAT64) AS ValorBruto,
            p.InfoExtra,
            ud.max_data AS data_referencia
        FROM {TABLE_POSICAO} p
        CROSS JOIN ultima_data ud
        LEFT JOIN clientes c ON SAFE_CAST(TRIM(p.Conta) AS INT64) = c.conta_num
        {join_part}
        WHERE DATE(p.Data) = ud.max_data
          AND p.Classe = 'Derivativo'
        ORDER BY p.Conta, SAFE_CAST(p.Vencimento AS DATE)
    """

    log.info(f"BQ opcoes: {'todos' if is_admin else assessor_name}")
    if qp:
        rows_raw = list(bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result())
    else:
        rows_raw = list(bq.query(sql).result())

    data_ref = None
    rows_list = []
    for row in rows_raw:
        d = dict(row)
        dr = d.pop("data_referencia", None)
        if data_ref is None and dr is not None:
            data_ref = dr.isoformat() if hasattr(dr, "isoformat") else str(dr)
        for field in ["Quantidade", "Preco", "ValorBruto"]:
            if d.get(field) is not None:
                try:
                    d[field] = float(d[field])
                except (TypeError, ValueError):
                    d[field] = None
        rows_list.append(d)

    # Cotações dos ativos subjacentes (ex: BOVA11, VALE3, ITUB4)
    refs = list({_ref_from_info(r.get("InfoExtra")) for r in rows_list} - {None})
    precos = buscar_precos(refs) if refs else {}

    resultado = {"data_referencia": data_ref, "opcoes": rows_list, "precos": precos}
    cache_set(cache_key, resultado)
    return resultado


@app.get("/api/posicao")
async def posicao(
    authorization: Optional[str] = Header(default=None),
):
    """
    Retorna AUC total e contagem de contas ativas para a data mais recente.
    Admin → totais do escritório. Assessor → apenas suas contas.
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    filter_assessor = None if is_admin else assessor_name

    cache_key = f"posicao:snapshot:{filter_assessor or 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    if filter_assessor:
        join_assessor = f"INNER JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ma.Conta"
        where_assessor = "AND UPPER(ma.Assessor) = UPPER(@assessor)"
        qp = [ScalarQueryParameter("assessor", "STRING", filter_assessor)]
    else:
        join_assessor  = ""
        where_assessor = ""
        qp = []

    sql = f"""
        WITH ultima_data AS (
            SELECT MAX(DATE(Data)) AS max_data
            FROM {TABLE_POSICAO}
        )
        SELECT
            ROUND(SUM(p.ValorBruto), 2)    AS auc_total,
            COUNT(DISTINCT p.Conta)         AS contas_ativas,
            ud.max_data                     AS data_referencia
        FROM {TABLE_POSICAO} p
        CROSS JOIN ultima_data ud
        {join_assessor}
        WHERE DATE(p.Data) = ud.max_data
          AND p.Classe != 'Aluguel de Ações'
          {where_assessor}
        GROUP BY ud.max_data
    """

    log.info(f"BQ posicao: assessor={filter_assessor or 'todos'}")
    if qp:
        rows = list(bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result())
    else:
        rows = list(bq.query(sql).result())

    if not rows:
        resultado = {"auc_total": 0, "contas_ativas": 0, "data_referencia": None}
    else:
        row = rows[0]
        data_ref = row["data_referencia"]
        resultado = {
            "auc_total":       float(row["auc_total"] or 0),
            "contas_ativas":   int(row["contas_ativas"] or 0),
            "data_referencia": data_ref.date().isoformat() if hasattr(data_ref, "date") else str(data_ref),
        }

    cache_set(cache_key, resultado)
    return resultado


# ── Modelo para criação de exceção ───────────────────────────────────────────
class ExcecaoBody(BaseModel):
    conta:       int
    assessor:    str
    data_inicio: str            # "YYYY-MM-DD"
    data_fim:    Optional[str] = None   # "YYYY-MM-DD" ou None
    motivo:      Optional[str] = None


@app.get("/api/config/excecoes")
async def listar_excecoes(
    authorization: Optional[str] = Header(default=None),
):
    """Lista todas as exceções de assessor — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    sql = f"""
        SELECT
            Conta,
            Assessor,
            FORMAT_DATE('%Y-%m-%d', DataInicio) AS DataInicio,
            FORMAT_DATE('%Y-%m-%d', DataFim)    AS DataFim,
            COALESCE(Motivo, '')                AS Motivo,
            FORMAT_DATETIME('%Y-%m-%dT%H:%M:%S', DataCriacao) AS DataCriacao,
            COALESCE(CriadoPor, '')             AS CriadoPor
        FROM {TABLE_EXCECOES}
        ORDER BY DataInicio DESC
    """
    rows = run_query(sql)
    return {"excecoes": rows}


@app.post("/api/config/excecoes")
async def criar_excecao(
    body: ExcecaoBody,
    authorization: Optional[str] = Header(default=None),
):
    """Cria uma nova exceção de assessor — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    email = token_data.get("email", "")

    data_fim_expr = f"DATE('{body.data_fim}')" if body.data_fim else "NULL"
    motivo_expr   = f"'{body.motivo.replace(chr(39), chr(39)*2)}'" if body.motivo else "NULL"
    assessor_esc  = body.assessor.replace("'", "''")

    sql = f"""
        INSERT INTO {TABLE_EXCECOES}
            (Conta, Assessor, DataInicio, DataFim, Motivo, DataCriacao, CriadoPor)
        VALUES
            ({body.conta}, '{assessor_esc}', DATE('{body.data_inicio}'),
             {data_fim_expr}, {motivo_expr},
             CURRENT_DATETIME(), '{email}')
    """
    try:
        bq.query(sql).result()
    except Exception as e:
        log.error(f"Erro ao criar exceção: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    # Invalida cache relacionado (exceções afetam posicoes, relatorio, assessores e opcoes)
    for k in list(_cache.keys()):
        if any(x in k for x in ("posicoes", "receitas", "relatorio", "assessores", "opcoes")):
            del _cache[k]

    return {"ok": True}


@app.delete("/api/config/excecoes")
async def deletar_excecao(
    conta:       int = Query(...),
    data_inicio: str = Query(...),   # "YYYY-MM-DD"
    authorization: Optional[str] = Header(default=None),
):
    """Remove uma exceção pelo par (Conta, DataInicio) — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    sql = f"""
        DELETE FROM {TABLE_EXCECOES}
        WHERE Conta = {conta}
          AND DataInicio = DATE('{data_inicio}')
    """
    try:
        bq.query(sql).result()
    except Exception as e:
        log.error(f"Erro ao deletar exceção: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    for k in list(_cache.keys()):
        if "posicoes" in k or "receitas" in k:
            del _cache[k]

    return {"ok": True}


# ── Pipeline: gestão de arquivos no GCS ──────────────────────────────────────

@app.get("/api/pipeline/arquivos")
async def listar_arquivos(
    authorization: Optional[str] = Header(default=None),
):
    """Lista arquivos na pasta entradas/ do bucket — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    bucket = gcs_client.bucket(GCS_BUCKET)
    blobs  = bucket.list_blobs(prefix=GCS_PREFIX)

    arquivos = []
    for blob in blobs:
        nome = blob.name.replace(GCS_PREFIX, "")
        if not nome:
            continue
        arquivos.append({
            "nome":        nome,
            "tamanho_kb":  round(blob.size / 1024, 1),
            "atualizado":  blob.updated.isoformat() if blob.updated else None,
        })

    arquivos.sort(key=lambda x: x["nome"])
    return {"arquivos": arquivos}


@app.post("/api/pipeline/upload")
async def upload_arquivo(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    """Faz upload de um arquivo para entradas/ no bucket — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos.")

    try:
        bucket = gcs_client.bucket(GCS_BUCKET)
        blob   = bucket.blob(f"{GCS_PREFIX}{file.filename}")
        blob.upload_from_file(file.file, content_type=file.content_type or "application/octet-stream")
        log.info(f"Upload GCS: {file.filename}")
        return {"ok": True, "nome": file.filename}
    except Exception as e:
        log.error(f"Erro upload GCS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/pipeline/arquivo")
async def deletar_arquivo(
    nome: str = Query(...),
    authorization: Optional[str] = Header(default=None),
):
    """Remove um arquivo de entradas/ no bucket — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    try:
        bucket = gcs_client.bucket(GCS_BUCKET)
        blob   = bucket.blob(f"{GCS_PREFIX}{nome}")
        blob.delete()
        log.info(f"Deletado GCS: {nome}")
        return {"ok": True}
    except Exception as e:
        log.error(f"Erro delete GCS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/pipeline/arquivo/download")
async def download_arquivo(
    nome: str = Query(...),
    authorization: Optional[str] = Header(default=None),
):
    """Faz download de um arquivo do bucket — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    try:
        bucket = gcs_client.bucket(GCS_BUCKET)
        blob   = bucket.blob(f"{GCS_PREFIX}{nome}")
        buffer = io.BytesIO()
        blob.download_to_file(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(nome)}"},
        )
    except Exception as e:
        log.error(f"Erro download GCS: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/excecoes/excel")
async def excecoes_excel(
    authorization: Optional[str] = Header(default=None),
):
    """Exporta todas as exceções como arquivo Excel — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    sql = f"""
        SELECT
            Conta,
            Assessor,
            FORMAT_DATE('%d/%m/%Y', DataInicio) AS DataInicio,
            FORMAT_DATE('%d/%m/%Y', DataFim)    AS DataFim,
            COALESCE(Motivo, '')                AS Motivo,
            COALESCE(CriadoPor, '')             AS CriadoPor,
            FORMAT_DATETIME('%d/%m/%Y %H:%M', DataCriacao) AS DataCriacao
        FROM {TABLE_EXCECOES}
        ORDER BY DataInicio DESC
    """
    rows = run_query(sql)
    df   = pd.DataFrame(rows)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Exceções")
        ws = writer.sheets["Exceções"]
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="excecoes_assessor.xlsx"'},
    )


# ── Relatórios ───────────────────────────────────────────────────────────────

@app.get("/api/assessores")
async def assessores_endpoint(
    authorization: Optional[str] = Header(default=None),
):
    """Lista assessores distintos — apenas admins (filtro do relatório)."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    cached = cache_get("assessores")
    if cached:
        return cached

    sql = f"""
        SELECT DISTINCT assessor
        FROM (
            SELECT UPPER(Assessor) AS assessor
            FROM {TABLE_ASSESSOR_BASE}
            WHERE MesRef = (SELECT MAX(MesRef) FROM {TABLE_ASSESSOR_BASE})
            UNION DISTINCT
            SELECT UPPER(Assessor) AS assessor
            FROM {TABLE_EXCECOES}
            WHERE DataInicio <= CURRENT_DATE()
              AND (DataFim IS NULL OR DataFim >= CURRENT_DATE())
        )
        WHERE assessor IS NOT NULL AND assessor != ''
        ORDER BY assessor
    """
    rows = run_query(sql)
    resultado = {"assessores": [r["assessor"] for r in rows]}
    cache_set("assessores", resultado)
    return resultado


@app.get("/api/relatorio/historico")
async def relatorio_historico(
    periodo: str = "12m",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    assessor: Optional[str] = None,
    authorization: Optional[str] = Header(default=None),
):
    """
    Relatório histórico: AuC diário, novas contas/mês, receita+ROA mensal,
    top 20 variações absolutas de PL (últimos 30 dias).

    periodo: "3m" | "6m" | "12m" | "all"
    assessor: nome do assessor (apenas admins podem especificar)
    """
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    filter_assessor = (assessor.strip() if assessor else None) if is_admin else assessor_name

    cache_key = f"relatorio:{periodo}:{start_date or ''}:{end_date or ''}:{filter_assessor or 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    mapa_periodo = {
        "3m":  "DATE_SUB(CURRENT_DATE(), INTERVAL 3  MONTH)",
        "6m":  "DATE_SUB(CURRENT_DATE(), INTERVAL 6  MONTH)",
        "12m": "DATE_SUB(CURRENT_DATE(), INTERVAL 12 MONTH)",
        "all": "DATE '2010-01-01'",
    }
    if periodo == "custom" and start_date:
        data_inicio = f"DATE '{start_date}'"
        data_fim    = f"DATE '{end_date}'" if end_date else "CURRENT_DATE()"
    else:
        data_inicio = mapa_periodo.get(periodo, mapa_periodo["12m"])
        data_fim    = "CURRENT_DATE()"

    qp = [ScalarQueryParameter("assessor", "STRING", filter_assessor)] if filter_assessor else []

    # CTEs que filtram posicao_das_contas por assessor (via conta_assessor_base + exceções)
    if filter_assessor:
        contas_cte = f"""
    contas_assessor AS (
        SELECT Conta AS conta_num
        FROM {_mapa_assessor_sq()}
        WHERE UPPER(Assessor) = UPPER(@assessor)
    ),"""
        contas_join          = "INNER JOIN contas_assessor ca ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ca.conta_num"
        contas_join_primeira = "INNER JOIN contas_assessor ca ON SAFE_CAST(pa.Conta AS INT64) = ca.conta_num"
        contas_join_p2       = "INNER JOIN contas_assessor ca ON SAFE_CAST(TRIM(p2.Conta) AS INT64) = ca.conta_num"
    else:
        contas_cte           = ""
        contas_join          = ""
        contas_join_primeira = ""
        contas_join_p2       = ""

    assessor_where = "AND UPPER(TRIM(Assessor_Manual)) = UPPER(@assessor)" if filter_assessor else ""

    def rq(sql: str) -> list[dict]:
        if qp:
            return [dict(r) for r in bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result()]
        return [dict(r) for r in bq.query(sql).result()]

    def ser(rows: list[dict]) -> list[dict]:
        return [{k: (v.isoformat() if hasattr(v, "isoformat") else v) for k, v in row.items()} for row in rows]

    # ── 1. AuC diário ─────────────────────────────────────────────────────────
    sql_auc = f"""
        WITH{contas_cte}
        auc_dia AS (
            SELECT DATE(p.Data) AS dia, ROUND(SUM(p.ValorBruto), 2) AS auc
            FROM {TABLE_POSICAO} p
            {contas_join}
            WHERE DATE(p.Data) >= {data_inicio}
              AND DATE(p.Data) <= {data_fim}
              AND p.Classe != 'Aluguel de Ações'
            GROUP BY DATE(p.Data)
        )
        SELECT FORMAT_DATE('%Y-%m-%d', dia) AS dia, auc
        FROM auc_dia
        ORDER BY dia
    """

    # ── 2. Novas contas e retiradas por mês ──────────────────────────────────────
    # Nova    = conta cuja DataPrimeiraAparicao (tabela pré-calculada) cai no período.
    # Retirada = conta cuja última aparição em posicao_das_contas cai no período,
    #            mas antes do mês mais recente da tabela (evita falso positivo).
    sql_novas = f"""
        WITH{contas_cte}
        max_mes AS (
            SELECT DATE_TRUNC(MAX(DATE(Data)), MONTH) AS mes_atual
            FROM {TABLE_POSICAO}
        ),
        novas AS (
            SELECT
                DATE_TRUNC(pa.DataPrimeiraAparicao, MONTH) AS mes,
                COUNT(*) AS novas_contas
            FROM {TABLE_PRIMEIRA_APARICAO} pa
            {contas_join_primeira}
            WHERE pa.DataPrimeiraAparicao >= {data_inicio}
              AND pa.DataPrimeiraAparicao <= {data_fim}
            GROUP BY 1
        ),
        ultima_aparicao AS (
            SELECT TRIM(p.Conta) AS Conta,
                   DATE_TRUNC(MAX(DATE(p.Data)), MONTH) AS mes_saida
            FROM {TABLE_POSICAO} p
            {contas_join}
            GROUP BY TRIM(p.Conta)
        ),
        retiradas AS (
            SELECT u.mes_saida AS mes, COUNT(*) AS contas_retiradas
            FROM ultima_aparicao u
            CROSS JOIN max_mes m
            WHERE u.mes_saida >= DATE_TRUNC({data_inicio}, MONTH)
              AND u.mes_saida <= DATE_TRUNC({data_fim},   MONTH)
              AND u.mes_saida <  m.mes_atual
            GROUP BY u.mes_saida
        )
        SELECT
            FORMAT_DATE('%Y-%m', COALESCE(n.mes, r.mes))   AS mes,
            COALESCE(n.novas_contas,     0)                 AS novas_contas,
            COALESCE(r.contas_retiradas, 0)                 AS contas_retiradas,
            COALESCE(n.novas_contas, 0) - COALESCE(r.contas_retiradas, 0) AS saldo_liquido
        FROM novas n
        FULL OUTER JOIN retiradas r ON n.mes = r.mes
        ORDER BY COALESCE(n.mes, r.mes)
    """

    # ── 3. Receita + ROA mensal ────────────────────────────────────────────────
    sql_receita = f"""
        WITH{contas_cte}
        receita_mensal AS (
            SELECT
                FORMAT_DATE('%Y-%m', DATE_TRUNC(DATE(Data_De_Referencia), MONTH)) AS mes,
                ROUND(SUM(Receita_Liquida),       2) AS receita_liquida,
                ROUND(SUM(Receita_Bruta),         2) AS receita_bruta,
                ROUND(SUM(Comissao_Liquida),      2) AS comissao_liquida,
                ROUND(SUM(Repasse_Total_liquido), 2) AS repasse_liquido
            FROM {TABLE}
            WHERE DATE(Data_De_Referencia) >= {data_inicio}
              AND DATE(Data_De_Referencia) <= {data_fim}
              {assessor_where}
            GROUP BY mes
        ),
        auc_mensal AS (
            SELECT
                FORMAT_DATE('%Y-%m', DATE_TRUNC(p.dia_data, MONTH)) AS mes,
                ROUND(AVG(p.auc_dia), 2) AS avg_auc
            FROM (
                SELECT DATE(p2.Data) AS dia_data, SUM(p2.ValorBruto) AS auc_dia
                FROM {TABLE_POSICAO} p2
                {contas_join_p2}
                WHERE DATE(p2.Data) >= {data_inicio}
                  AND DATE(p2.Data) <= {data_fim}
                  AND p2.Classe != 'Aluguel de Ações'
                GROUP BY DATE(p2.Data)
            ) p
            GROUP BY mes
        )
        SELECT
            r.mes,
            r.receita_liquida,
            r.receita_bruta,
            r.comissao_liquida,
            r.repasse_liquido,
            ROUND(COALESCE(a.avg_auc, 0), 0) AS avg_auc,
            CASE WHEN COALESCE(a.avg_auc, 0) > 0
                 THEN ROUND((r.receita_bruta / a.avg_auc) * 12 * 100, 4)
                 ELSE NULL
            END AS roa_anualizado_pct
        FROM receita_mensal r
        LEFT JOIN auc_mensal a USING (mes)
        ORDER BY r.mes
    """

    # ── 4. Top 20 PL movers (delta AuC vs ~30 dias atrás) ────────────────────
    sql_movers = f"""
        WITH ultima_data AS (
            SELECT MAX(DATE(Data)) AS max_data
            FROM {TABLE_POSICAO}
        ),
        data_ref AS (
            SELECT MAX(DATE(p.Data)) AS ref_data
            FROM {TABLE_POSICAO} p
            CROSS JOIN ultima_data ud
            WHERE DATE(p.Data) <= DATE_SUB(ud.max_data, INTERVAL 30 DAY)
        ),{contas_cte}
        auc_atual AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto), 2) AS auc
            FROM {TABLE_POSICAO} p
            CROSS JOIN ultima_data ud
            {contas_join}
            WHERE DATE(p.Data) = ud.max_data
              AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        auc_ant AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto), 2) AS auc
            FROM {TABLE_POSICAO} p
            CROSS JOIN data_ref dr
            {contas_join}
            WHERE DATE(p.Data) = dr.ref_data
              AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE}
            WHERE Cliente IS NOT NULL AND Cliente != ''
            GROUP BY Conta
        )
        SELECT
            a.Conta,
            COALESCE(n.cliente,   '') AS cliente,
            UPPER(COALESCE(ma.Assessor, '')) AS assessor,
            COALESCE(a.auc, 0)        AS auc_atual,
            COALESCE(r.auc, 0)        AS auc_ref,
            ROUND(COALESCE(a.auc, 0) - COALESCE(r.auc, 0), 2) AS delta_auc
        FROM auc_atual a
        LEFT JOIN auc_ant r ON a.Conta = r.Conta
        LEFT JOIN nomes n   ON SAFE_CAST(a.Conta AS INT64) = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(a.Conta AS INT64) = ma.Conta
        ORDER BY ABS(COALESCE(a.auc, 0) - COALESCE(r.auc, 0)) DESC
        LIMIT 20
    """

    # ── 5. Top 20 Categoria × Cliente por Receita Bruta ─────────────────────
    sql_top_sub = f"""
        SELECT
            COALESCE(NULLIF(TRIM(Categoria), ''), 'Sem Categoria') AS subclasse,
            COALESCE(NULLIF(TRIM(Produto),   ''), '—')             AS produto,
            CAST(Conta AS STRING)                                   AS Conta,
            COALESCE(NULLIF(TRIM(Cliente), ''), '—')               AS cliente,
            UPPER(COALESCE(TRIM(Assessor_Manual), ''))             AS assessor,
            ROUND(SUM(Receita_Bruta),         2) AS receita_bruta,
            ROUND(SUM(Receita_Liquida),       2) AS receita_liquida,
            ROUND(SUM(Comissao_Liquida),      2) AS comissao_liquida,
            ROUND(SUM(Repasse_Total_liquido), 2) AS repasse
        FROM {TABLE}
        WHERE DATE(Data_De_Referencia) >= {data_inicio}
          AND DATE(Data_De_Referencia) <= {data_fim}
          {assessor_where}
          AND Categoria IS NOT NULL AND TRIM(Categoria) != ''
        GROUP BY subclasse, produto, Conta, cliente, assessor
        ORDER BY receita_bruta DESC
        LIMIT 20
    """

    log.info(f"BQ relatorio/historico: periodo={periodo}, assessor={filter_assessor or 'todos'}")
    auc_diario   = ser(rq(sql_auc))
    novas_contas = ser(rq(sql_novas))
    receita_roa  = ser(rq(sql_receita))
    pl_movers    = ser(rq(sql_movers))
    top_subclass = ser(rq(sql_top_sub))

    # KPIs derivados
    auc_atual_val     = float(auc_diario[-1]["auc"]) if auc_diario else 0
    delta_auc         = round(
        float(auc_diario[-1]["auc"]) - float(auc_diario[0]["auc"]), 2
    ) if len(auc_diario) >= 2 else 0
    novas_total       = sum(int(r.get("novas_contas") or 0) for r in novas_contas)
    receita_total     = round(sum(float(r.get("receita_liquida") or 0) for r in receita_roa), 2)
    roa_ultimo        = float(receita_roa[-1]["roa_anualizado_pct"]) if receita_roa and receita_roa[-1].get("roa_anualizado_pct") is not None else None

    resultado = {
        "periodo":      periodo,
        "assessor":     filter_assessor,
        "gerado_em":    datetime.utcnow().isoformat(),
        "kpis": {
            "auc_atual":            auc_atual_val,
            "delta_auc_periodo":    delta_auc,
            "novas_contas_periodo": novas_total,
            "roa_anualizado_pct":   roa_ultimo,
            "receita_periodo":      receita_total,
        },
        "auc_diario":   auc_diario,
        "novas_contas": novas_contas,
        "receita_roa":  receita_roa,
        "pl_movers":    pl_movers,
        "top_subclass": top_subclass,
    }

    cache_set(cache_key, resultado)
    return resultado


@app.delete("/api/cache")
async def limpar_cache(
    authorization: Optional[str] = Header(default=None),
):
    """Limpa o cache manualmente — apenas admins. Útil após carga de novos dados."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin" and token_data.get("email") not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Apenas administradores podem limpar o cache.")
    _cache.clear()
    return {"ok": True, "message": "Cache limpo."}


@app.post("/api/cache/clear")
async def limpar_cache_pipeline(x_pipeline_key: Optional[str] = Header(default=None)):
    """Limpa cache via API key — chamado pelo pipeline Python após carga de novos dados."""
    expected = os.getenv("PIPELINE_API_KEY", "")
    if not expected or x_pipeline_key != expected:
        raise HTTPException(status_code=403, detail="Chave inválida.")
    _cache.clear()
    log.info("Cache limpo via pipeline API key.")
    return {"ok": True, "mensagem": f"Cache limpo. {len(_cache)} entradas removidas."}


# ══════════════════════════════════════════════════════════════════════════════
# /api/webhook/btg — Recebe payloads do BTG, baixa arquivo S3 e salva no GCS
# ══════════════════════════════════════════════════════════════════════════════
@app.post("/api/webhook/btg")
async def webhook_btg(request: Request, token: Optional[str] = Query(default=None)):
    """
    Recebe webhooks do BTG Pactual (cadastrados no portal de desenvolvedor).
    Valida o token, grava o payload no BigQuery, baixa o arquivo S3 imediatamente
    e salva no GCS antes que a URL assinada expire (TTL = 1h).

    Tipos tratados:
    - account-advisor  → gs://creta-btg-pipeline/entradas/
    - partner-report   → gs://creta-btg-pipeline/carteira-recomendada/
    """
    # ── 1. Valida token ───────────────────────────────────────────────────────
    expected_token = os.getenv("WEBHOOK_TOKEN", "creta-btg-webhook-2024")
    if token != expected_token:
        log.warning("Webhook BTG: token inválido recebido.")
        raise HTTPException(status_code=403, detail="Token inválido.")

    # ── 2. Lê payload ─────────────────────────────────────────────────────────
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Payload JSON inválido.")

    received_at = datetime.utcnow().isoformat() + "Z"
    log.info(f"Webhook BTG recebido: keys={list(payload.keys())}")

    # ── 3. Grava raw no BigQuery ──────────────────────────────────────────────
    try:
        erros_bq = bq.insert_rows_json(
            TABLE_WEBHOOK_RAW,
            [{"payload": json.dumps(payload, ensure_ascii=False), "received_at": received_at}],
        )
        if erros_bq:
            log.error(f"Webhook: erro ao gravar BigQuery: {erros_bq}")
    except Exception as e:
        log.error(f"Webhook: exceção ao gravar BigQuery: {e}")

    # ── 4. Extrai URL S3 ──────────────────────────────────────────────────────
    # Formato A (account-advisor):  {"url": "...", "fileSize": ..., ...}
    # Formato B (partner-report):   {"errors": [], "response": {"url": "...", ...}}
    erros_payload = payload.get("errors") or []
    if erros_payload:
        # BTG reportou erro — sem arquivo para baixar
        log.warning(f"Webhook BTG com erros: {erros_payload}")
        return {"ok": True, "arquivo": None, "aviso": "Payload com erros — sem arquivo."}

    resp   = payload.get("response") or {}
    s3_url = payload.get("url") or resp.get("url")

    if not s3_url:
        log.info("Webhook BTG sem URL S3 — nenhum arquivo para baixar.")
        return {"ok": True, "arquivo": None}

    # ── 5. Determina pasta GCS e nome do arquivo ──────────────────────────────
    if "account-advisor" in s3_url:
        gcs_folder = "entradas/"
    else:
        gcs_folder = "carteira-recomendada/"

    # Pega o nome do arquivo da URL (antes dos query params)
    filename = s3_url.split("?")[0].split("/")[-1]
    gcs_path = gcs_folder + filename

    # ── 6. Baixa arquivo S3 e salva no GCS ───────────────────────────────────
    try:
        log.info(f"Webhook: baixando {filename} ...")
        r = requests.get(s3_url, timeout=30)
        r.raise_for_status()

        bucket = gcs_client.bucket(GCS_BUCKET)
        blob   = bucket.blob(gcs_path)
        blob.upload_from_string(r.content, content_type="application/octet-stream")

        log.info(f"Webhook: salvo em gs://{GCS_BUCKET}/{gcs_path} ({len(r.content)} bytes)")
    except Exception as e:
        # Não levanta erro para o BTG não retentar — o payload já está no BQ
        log.error(f"Webhook: falha ao baixar/salvar arquivo: {e}")
        return {"ok": True, "arquivo": None, "aviso": str(e)}

    return {"ok": True, "arquivo": f"gs://{GCS_BUCKET}/{gcs_path}"}


# ══════════════════════════════════════════════════════════════════════════════
# /api/gestao — Oportunidades e gestão de carteira para assessores
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/api/gestao")
async def gestao_carteira(
    assessor:      Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
):
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_name = token_data.get("assessor_name")
    is_admin      = role == "admin"

    filter_assessor = (assessor.strip() if assessor else None) if is_admin else assessor_name

    cache_key = f"gestao:{filter_assessor or 'all'}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    qp = [ScalarQueryParameter("assessor", "STRING", filter_assessor)] if filter_assessor else []

    def rq(sql: str) -> list[dict]:
        def ser_row(row):
            return {k: (v.isoformat() if hasattr(v, "isoformat") else v) for k, v in dict(row).items()}
        if qp:
            return [ser_row(r) for r in bq.query(sql, job_config=QueryJobConfig(query_parameters=qp)).result()]
        return [ser_row(r) for r in bq.query(sql).result()]

    # CTEs de filtro por assessor (via conta_assessor_base + exceções)
    if filter_assessor:
        contas_cte  = f"""contas_ass AS (
            SELECT Conta AS conta_num FROM {_mapa_assessor_sq()}
            WHERE UPPER(Assessor) = UPPER(@assessor)
        ),"""
        contas_join_p = "INNER JOIN contas_ass ca ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ca.conta_num"
        contas_join   = "INNER JOIN contas_ass ca ON SAFE_CAST(TRIM(Conta) AS INT64) = ca.conta_num"
    else:
        contas_cte    = ""
        contas_join_p = ""
        contas_join   = ""

    # ── 1. Caixa Parado ───────────────────────────────────────────────────────
    sql_caixa = f"""
        WITH ultima_data AS (SELECT MAX(DATE(Data)) AS d FROM {TABLE_POSICAO}),
        {contas_cte}
        posicao_total AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto),2) AS total_auc
            FROM {TABLE_POSICAO} p CROSS JOIN ultima_data ud
            {contas_join_p}
            WHERE DATE(p.Data) = ud.d AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        caixa AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto),2) AS valor_caixa
            FROM {TABLE_POSICAO} p CROSS JOIN ultima_data ud
            {contas_join_p}
            WHERE DATE(p.Data) = ud.d AND p.Classe = 'Caixa'
            GROUP BY TRIM(p.Conta)
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE} WHERE Cliente IS NOT NULL AND Cliente != '' GROUP BY Conta
        )
        SELECT
            c.Conta,
            COALESCE(n.cliente, '') AS cliente,
            UPPER(COALESCE(ma.Assessor,'')) AS assessor,
            c.valor_caixa,
            t.total_auc,
            CASE WHEN t.total_auc > 0 THEN ROUND(c.valor_caixa / t.total_auc * 100, 1) ELSE NULL END AS pct_caixa
        FROM caixa c
        LEFT JOIN posicao_total t ON c.Conta = t.Conta
        LEFT JOIN nomes n ON SAFE_CAST(c.Conta AS INT64) = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(c.Conta AS INT64) = ma.Conta
        WHERE c.valor_caixa >= 5000
        ORDER BY c.valor_caixa DESC
        LIMIT 50
    """

    # ── 2. Vencimentos nos próximos 90 dias ───────────────────────────────────
    sql_venc = f"""
        WITH ultima_data AS (SELECT MAX(DATE(Data)) AS d FROM {TABLE_POSICAO}),
        {contas_cte}
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE} WHERE Cliente IS NOT NULL AND Cliente != '' GROUP BY Conta
        )
        SELECT
            TRIM(p.Conta) AS Conta,
            COALESCE(n.cliente,'') AS cliente,
            UPPER(COALESCE(ma.Assessor,'')) AS assessor,
            p.Classe, p.Subclasse, p.Nome,
            p.Vencimento,
            DATE_DIFF(DATE(p.Vencimento), CURRENT_DATE(), DAY) AS dias_para_vencer,
            ROUND(SUM(p.ValorBruto),2) AS valor
        FROM {TABLE_POSICAO} p
        CROSS JOIN ultima_data ud
        {contas_join_p}
        LEFT JOIN nomes n ON SAFE_CAST(TRIM(p.Conta) AS INT64) = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(TRIM(p.Conta) AS INT64) = ma.Conta
        WHERE DATE(p.Data) = ud.d
          AND p.Vencimento IS NOT NULL AND p.Vencimento != ''
          AND DATE(p.Vencimento) >= CURRENT_DATE()
          AND DATE(p.Vencimento) <= DATE_ADD(CURRENT_DATE(), INTERVAL 90 DAY)
          AND p.Classe NOT IN ('Ação','Caixa','Aluguel de Ações','Derivativo')
        GROUP BY TRIM(p.Conta), cliente, assessor, p.Classe, p.Subclasse, p.Nome, p.Vencimento
        ORDER BY p.Vencimento
        LIMIT 100
    """

    # ── 3. ROA por Cliente (últimos 12 meses) ─────────────────────────────────
    sql_roa = f"""
        WITH {contas_cte}
        receita_conta AS (
            SELECT Conta,
                   ROUND(SUM(Receita_Bruta),2) AS receita_bruta
            FROM {TABLE}
            WHERE DATE(Data_De_Referencia) >= DATE_SUB(CURRENT_DATE(), INTERVAL 12 MONTH)
              AND UPPER(COALESCE(Assessor_Manual,'')) != ''
              {'AND UPPER(TRIM(Assessor_Manual)) = UPPER(@assessor)' if filter_assessor else ''}
            GROUP BY Conta
        ),
        auc_medio AS (
            SELECT p.conta_num,
                   ROUND(AVG(p.auc_dia),2) AS avg_auc
            FROM (
                SELECT DATE(p2.Data) AS d,
                       SAFE_CAST(TRIM(p2.Conta) AS INT64) AS conta_num,
                       SUM(p2.ValorBruto) AS auc_dia
                FROM {TABLE_POSICAO} p2
                {contas_join_p.replace('p.', 'p2.')}
                WHERE DATE(p2.Data) >= DATE_SUB(CURRENT_DATE(), INTERVAL 12 MONTH)
                  AND p2.Classe != 'Aluguel de Ações'
                GROUP BY d, SAFE_CAST(TRIM(p2.Conta) AS INT64)
            ) p
            GROUP BY p.conta_num
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE} WHERE Cliente IS NOT NULL AND Cliente != '' GROUP BY Conta
        )
        SELECT
            r.Conta,
            COALESCE(n.cliente,'') AS cliente,
            UPPER(COALESCE(ma.Assessor,'')) AS assessor,
            r.receita_bruta,
            ROUND(COALESCE(a.avg_auc,0),0) AS avg_auc,
            CASE WHEN COALESCE(a.avg_auc,0) > 0
                 THEN ROUND(r.receita_bruta / a.avg_auc * 12 * 100, 4) ELSE NULL END AS roa_pct
        FROM receita_conta r
        LEFT JOIN auc_medio a ON r.Conta = a.conta_num
        LEFT JOIN nomes n ON r.Conta = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON r.Conta = ma.Conta
        WHERE COALESCE(a.avg_auc,0) > 0
        ORDER BY roa_pct ASC
        LIMIT 50
    """

    # ── 4. Clientes sem receita há 60+ dias ───────────────────────────────────
    sql_sem_rec = f"""
        WITH ultima_data AS (SELECT MAX(DATE(Data)) AS d FROM {TABLE_POSICAO}),
        {contas_cte}
        contas_ativas AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto),2) AS auc_atual
            FROM {TABLE_POSICAO} p CROSS JOIN ultima_data ud
            {contas_join_p}
            WHERE DATE(p.Data) = ud.d AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        ultima_receita AS (
            SELECT Conta, MAX(DATE(Data_De_Referencia)) AS ultima_data
            FROM {TABLE}
            GROUP BY Conta
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE} WHERE Cliente IS NOT NULL AND Cliente != '' GROUP BY Conta
        )
        SELECT
            ca.Conta,
            COALESCE(n.cliente,'') AS cliente,
            UPPER(COALESCE(ma.Assessor,'')) AS assessor,
            ca.auc_atual,
            ur.ultima_data AS ultima_receita,
            DATE_DIFF(CURRENT_DATE(), COALESCE(ur.ultima_data, DATE('2020-01-01')), DAY) AS dias_sem_receita
        FROM contas_ativas ca
        LEFT JOIN ultima_receita ur ON SAFE_CAST(ca.Conta AS INT64) = ur.Conta
        LEFT JOIN nomes n ON SAFE_CAST(ca.Conta AS INT64) = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(ca.Conta AS INT64) = ma.Conta
        WHERE (ur.ultima_data IS NULL OR ur.ultima_data < DATE_SUB(CURRENT_DATE(), INTERVAL 60 DAY))
          AND ca.auc_atual > 1000
        ORDER BY dias_sem_receita DESC
        LIMIT 50
    """

    # ── 5. AuC em queda (vs 60 dias atrás) ───────────────────────────────────
    sql_queda = f"""
        WITH ultima_data AS (SELECT MAX(DATE(Data)) AS d FROM {TABLE_POSICAO}),
        data_60 AS (
            SELECT MAX(DATE(p.Data)) AS d
            FROM {TABLE_POSICAO} p CROSS JOIN ultima_data ud
            WHERE DATE(p.Data) <= DATE_SUB(ud.d, INTERVAL 60 DAY)
        ),
        {contas_cte}
        auc_atual AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto),2) AS auc
            FROM {TABLE_POSICAO} p CROSS JOIN ultima_data ud
            {contas_join_p}
            WHERE DATE(p.Data) = ud.d AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        auc_ant AS (
            SELECT TRIM(p.Conta) AS Conta, ROUND(SUM(p.ValorBruto),2) AS auc
            FROM {TABLE_POSICAO} p CROSS JOIN data_60 d60
            {contas_join_p}
            WHERE DATE(p.Data) = d60.d AND p.Classe != 'Aluguel de Ações'
            GROUP BY TRIM(p.Conta)
        ),
        nomes AS (
            SELECT Conta AS conta_num, MAX(Cliente) AS cliente
            FROM {TABLE} WHERE Cliente IS NOT NULL AND Cliente != '' GROUP BY Conta
        )
        SELECT
            a.Conta,
            COALESCE(n.cliente,'') AS cliente,
            UPPER(COALESCE(ma.Assessor,'')) AS assessor,
            a.auc AS auc_atual,
            r.auc AS auc_60d,
            ROUND(a.auc - r.auc, 2) AS delta,
            CASE WHEN r.auc > 0 THEN ROUND((a.auc - r.auc) / r.auc * 100, 1) ELSE NULL END AS pct_var
        FROM auc_atual a
        INNER JOIN auc_ant r ON a.Conta = r.Conta
        LEFT JOIN nomes n ON SAFE_CAST(a.Conta AS INT64) = n.conta_num
        LEFT JOIN {_mapa_assessor_sq()} ma ON SAFE_CAST(a.Conta AS INT64) = ma.Conta
        WHERE a.auc < r.auc
        ORDER BY delta ASC
        LIMIT 50
    """

    log.info(f"BQ gestao: assessor={filter_assessor or 'todos'}")
    caixa_parado  = rq(sql_caixa)
    vencimentos   = rq(sql_venc)
    roa_cliente   = rq(sql_roa)
    sem_receita   = rq(sql_sem_rec)
    auc_queda     = rq(sql_queda)

    # ── 6. Lista de prioridades (score combinado) ─────────────────────────────
    scores: dict[str, dict] = {}
    for r in caixa_parado:
        c = r["Conta"]
        scores.setdefault(c, {"Conta": c, "cliente": r["cliente"], "assessor": r["assessor"], "score": 0, "alertas": []})
        scores[c]["score"] += 3
        scores[c]["alertas"].append(f"Caixa R$ {fmt_brl(r['valor_caixa'])}")
        scores[c]["caixa"] = r["valor_caixa"]
    # Agrupar vencimentos por conta: pegar o mais próximo + contar total
    _venc_by_conta: dict[str, dict] = {}
    for r in vencimentos:
        c = r["Conta"]
        d = r["dias_para_vencer"]
        if c not in _venc_by_conta:
            _venc_by_conta[c] = {"min_dias": d, "count": 1, "cliente": r["cliente"], "assessor": r["assessor"]}
        else:
            _venc_by_conta[c]["count"] += 1
            if d < _venc_by_conta[c]["min_dias"]:
                _venc_by_conta[c]["min_dias"] = d
    for c, v in _venc_by_conta.items():
        scores.setdefault(c, {"Conta": c, "cliente": v["cliente"], "assessor": v["assessor"], "score": 0, "alertas": []})
        scores[c]["score"] += 2
        cnt = f" ({v['count']})" if v["count"] > 1 else ""
        scores[c]["alertas"].append(f"Vence em {v['min_dias']}d{cnt}")
    for r in sem_receita:
        c = r["Conta"]
        scores.setdefault(c, {"Conta": c, "cliente": r["cliente"], "assessor": r["assessor"], "score": 0, "alertas": []})
        scores[c]["score"] += 2
        scores[c]["alertas"].append(f"Sem receita há {r['dias_sem_receita']}d")
    for r in auc_queda:
        c = r["Conta"]
        scores.setdefault(c, {"Conta": c, "cliente": r["cliente"], "assessor": r["assessor"], "score": 0, "alertas": []})
        scores[c]["score"] += 1
        scores[c]["alertas"].append(f"AuC -R$ {fmt_brl(r['delta'])}")

    prioridades = sorted(scores.values(), key=lambda x: x["score"], reverse=True)[:20]
    for p in prioridades:
        p["alertas"] = " | ".join(p["alertas"])

    resultado = {
        "gerado_em":    __import__("datetime").datetime.utcnow().isoformat(),
        "assessor":     filter_assessor,
        "caixa_parado": caixa_parado,
        "vencimentos":  vencimentos,
        "roa_cliente":  roa_cliente,
        "sem_receita":  sem_receita,
        "auc_queda":    auc_queda,
        "prioridades":  prioridades,
    }
    cache_set(cache_key, resultado)
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# /api/produtos — Carteira recomendada de ações (allocation + portfolio)
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/api/produtos")
async def produtos(authorization: Optional[str] = Header(default=None)):
    """Retorna carteira recomendada de equities (allocation + portfolio detalhado)."""
    await verificar_token(authorization)

    cache_key = "produtos:carteira_recomendada"
    cached = cache_get(cache_key)
    if cached:
        return cached

    def ser(rows):
        return [{k: (v.isoformat() if hasattr(v, "isoformat") else v) for k, v in dict(r).items()} for r in rows]

    sql_allocation = f"""
        SELECT Carteira, Tipo, Inicio, Fim,
               Rentab_Anterior, Rentab_Acumulada,
               Indice_Anterior, Indice_Acumulado,
               Ticker, Empresa, Setor, Peso
        FROM {TABLE_CR_ALLOCATION}
        WHERE DataExtracao = (SELECT MAX(DataExtracao) FROM {TABLE_CR_ALLOCATION})
        ORDER BY Carteira, Peso DESC
    """

    sql_portfolio = f"""
        SELECT Carteira, Benchmark, Indice_Acumulado, Rentab_Acumulada,
               Empresa, Setor, EV_EBITDA, PL, PVP
        FROM {TABLE_CR_PORTFOLIO}
        WHERE DataExtracao = (SELECT MAX(DataExtracao) FROM {TABLE_CR_PORTFOLIO})
        ORDER BY Carteira, Empresa
    """

    try:
        allocation = ser(bq.query(sql_allocation).result())
        portfolio  = ser(bq.query(sql_portfolio).result())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    resultado = {
        "allocation": allocation,
        "portfolio":  portfolio,
    }
    cache_set(cache_key, resultado)
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# /api/renda-fixa — CDB / LCA / LCI / LF (partner_report_cdb_lca_lci_lf)
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/renda-fixa")
async def renda_fixa(authorization: Optional[str] = Header(default=None)):
    """Retorna produtos de renda fixa (CDB/LCA/LCI/LF) da última extração."""
    await verificar_token(authorization)

    cache_key = "renda-fixa"
    cached = cache_get(cache_key)
    if cached:
        return cached

    sql = f"""
        SELECT
            productID,
            productName,
            issuerName,
            riskLevel,
            riskName,
            indexCaptureName,
            percentIndexValue,
            taxValue,
            cdbCdiIndexEquivalent,
            cdbYearIndexEquivalent,
            typeInterests,
            typeAmortization,
            typeLiquidityName,
            minAplicationValue,
            maxApplicationValue,
            amountTimeMonth,
            amountTimeDay,
            puValue,
            availableBallast,
            ballastAvailableValue,
            incomeTaxFree,
            secondary,
            isHighlighted,
            applicationDate,
            applicationDeadline,
            settlementDate,
            descriptionTimeLimit,
            DataExtracao
        FROM {TABLE_CDB_LCA}
        WHERE DataExtracao = (SELECT MAX(DataExtracao) FROM {TABLE_CDB_LCA})
        ORDER BY productName, riskLevel, taxValue DESC
    """

    try:
        rows = [{k: (v.isoformat() if hasattr(v, "isoformat") else v) for k, v in dict(r).items()}
                for r in bq.query(sql).result()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    data_extracao = rows[0]["DataExtracao"] if rows else None
    resultado = {"cdb_lca": rows, "data_extracao": data_extracao, "total": len(rows)}
    cache_set(cache_key, resultado)
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# /api/comite/recomendacoes — Recomendações do Comitê de Produtos
# Firestore: collection "comite", document "recomendacoes"
# Estrutura: { "recomendadas": ["Nome Carteira A", "Nome Carteira B"] }
# ══════════════════════════════════════════════════════════════════════════════

class ComiteRecomendacaoBody(BaseModel):
    carteira: str
    recomendado: bool


@app.get("/api/comite/recomendacoes")
async def get_comite_recomendacoes(
    authorization: Optional[str] = Header(default=None),
):
    """Retorna lista de carteiras recomendadas pelo comitê."""
    await verificar_token(authorization)
    try:
        doc = get_fs().collection("comite").document("recomendacoes").get()
        data = doc.to_dict() or {}
        return {"recomendadas": data.get("recomendadas", [])}
    except Exception as e:
        log.error(f"Firestore get recomendacoes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/comite/recomendacoes")
async def set_comite_recomendacao(
    body: ComiteRecomendacaoBody,
    authorization: Optional[str] = Header(default=None),
):
    """Marca ou desmarca uma carteira como recomendada pelo comitê — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")

    try:
        ref = get_fs().collection("comite").document("recomendacoes")
        doc = ref.get()
        data = doc.to_dict() or {}

        carteiras = set(data.get("recomendadas", []))
        if body.recomendado:
            carteiras.add(body.carteira)
        else:
            carteiras.discard(body.carteira)

        ref.set({"recomendadas": sorted(carteiras)})
        log.info(f"Comitê: {body.carteira} → recomendado={body.recomendado}")
        return {"ok": True, "carteira": body.carteira, "recomendado": body.recomendado}
    except Exception as e:
        log.error(f"Firestore set recomendacao: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Comitê RF: recomendações de Renda Fixa (por productID) ───────────────────
# Firestore: collection "comite", document "recomendacoes_rf"
# Estrutura: { "recomendadas": ["productID1", "productID2", ...] }

class ComiteRFBody(BaseModel):
    product_id: str
    recomendado: bool


@app.get("/api/comite/recomendacoes-rf")
async def get_comite_rf(authorization: Optional[str] = Header(default=None)):
    """Retorna IDs dos produtos de renda fixa recomendados pelo comitê."""
    await verificar_token(authorization)
    try:
        doc = get_fs().collection("comite").document("recomendacoes_rf").get()
        data = doc.to_dict() or {}
        return {"recomendadas": data.get("recomendadas", [])}
    except Exception as e:
        log.error(f"Firestore get recomendacoes_rf: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/comite/recomendacoes-rf")
async def set_comite_rf(
    body: ComiteRFBody,
    authorization: Optional[str] = Header(default=None),
):
    """Marca ou desmarca um produto de RF como recomendado pelo comitê — apenas admins."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Apenas administradores.")
    try:
        ref = get_fs().collection("comite").document("recomendacoes_rf")
        doc = ref.get()
        data = doc.to_dict() or {}
        ids = set(data.get("recomendadas", []))
        if body.recomendado:
            ids.add(body.product_id)
        else:
            ids.discard(body.product_id)
        ref.set({"recomendadas": sorted(ids)})
        log.info(f"Comitê RF: {body.product_id} → recomendado={body.recomendado}")
        return {"ok": True, "product_id": body.product_id, "recomendado": body.recomendado}
    except Exception as e:
        log.error(f"Firestore set recomendacoes_rf: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════════════════════════════════════
# /api/pipeline — Pipeline de clientes (Firestore: pipeline_clientes)
# Admin vê todos; assessor vê apenas os próprios.
# ══════════════════════════════════════════════════════════════════════════════

class PipelineCliente(BaseModel):
    nome:               str
    conta_antiga:       Optional[str]  = None
    origem:             Optional[str]  = None
    telefone:           Optional[str]  = None
    email:              Optional[str]  = None
    auc_global:         Optional[float] = None
    rf:                 Optional[float] = None
    rv:                 Optional[float] = None
    fundos:             Optional[float] = None
    prev:               Optional[float] = None
    coe:                Optional[float] = None
    ranking:            Optional[str]  = None   # Com certeza | Possivelmente | Stand By | Desafio | Não Vem
    pipe:               Optional[str]  = None   # Quente | Morno | Frio | LEAD
    primeiro_contato:   Optional[str]  = None   # ISO date
    retorno_positivo:   Optional[str]  = None   # Sim | Não
    status:             Optional[str]  = None
    conta_btg:          Optional[str]  = None
    pipe_quente:        Optional[float] = None
    auc_btg:            Optional[float] = None
    plano_acao:         Optional[str]  = None
    cliente_revertido:  Optional[str]  = None   # Sim | Não
    produtos:           Optional[list]  = None  # até 5 strings
    ultimo_contato:     Optional[str]  = None   # ISO date
    proximo_fup:        Optional[str]  = None   # ISO date
    observacoes:        Optional[str]  = None


def _pipeline_col():
    return get_fs().collection("pipeline_clientes")


def _ser_pipeline(doc) -> dict:
    d = doc.to_dict() or {}
    d["id"] = doc.id
    # Converte timestamps Firestore → ISO string
    for k, v in d.items():
        if hasattr(v, "isoformat"):
            d[k] = v.isoformat()
    return d


@app.get("/api/pipeline")
async def pipeline_listar(
    authorization: Optional[str] = Header(default=None),
):
    """Lista clientes do pipeline. Admin vê todos; assessor vê só os seus."""
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_uid  = token_data.get("uid")
    assessor_name = token_data.get("assessor_name", "")

    try:
        col = _pipeline_col()
        if role == "admin":
            docs = col.stream()
        else:
            docs = col.where("assessor_uid", "==", assessor_uid).stream()
        clientes = [_ser_pipeline(d) for d in docs]
        return {"clientes": clientes, "total": len(clientes)}
    except Exception as e:
        log.error(f"pipeline_listar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/pipeline")
async def pipeline_criar(
    body: PipelineCliente,
    authorization: Optional[str] = Header(default=None),
):
    """Cria novo cliente no pipeline."""
    token_data    = await verificar_token(authorization)
    assessor_uid  = token_data.get("uid")
    assessor_name = token_data.get("assessor_name", "") or token_data.get("email", "")

    dados = body.dict()
    dados["assessor_uid"]  = assessor_uid
    dados["assessor_name"] = assessor_name
    dados["created_at"]    = datetime.utcnow().isoformat()
    dados["updated_at"]    = datetime.utcnow().isoformat()

    # AuC potencial calculado
    auc_global = dados.get("auc_global") or 0
    auc_btg    = dados.get("auc_btg")    or 0
    dados["auc_potencial"] = round(auc_global - auc_btg, 2)

    try:
        ref = _pipeline_col().add(dados)
        return {"ok": True, "id": ref[1].id}
    except Exception as e:
        log.error(f"pipeline_criar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/pipeline/{doc_id}")
async def pipeline_atualizar(
    doc_id: str,
    body: PipelineCliente,
    authorization: Optional[str] = Header(default=None),
):
    """Atualiza cliente do pipeline — apenas o dono ou admin."""
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")

    try:
        ref = _pipeline_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")

        dados = {k: v for k, v in body.dict().items() if v is not None}
        dados["updated_at"] = datetime.utcnow().isoformat()

        auc_global = body.auc_global or doc.to_dict().get("auc_global") or 0
        auc_btg    = body.auc_btg    or doc.to_dict().get("auc_btg")    or 0
        dados["auc_potencial"] = round(auc_global - auc_btg, 2)

        ref.update(dados)
        return {"ok": True, "id": doc_id}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"pipeline_atualizar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/pipeline/{doc_id}")
async def pipeline_deletar(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    """Remove cliente do pipeline — apenas o dono ou admin."""
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")

    try:
        ref = _pipeline_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")
        ref.delete()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"pipeline_deletar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Pipeline: log de contatos ─────────────────────────────────────────────────

class ContatoLog(BaseModel):
    data:        str            # ISO date YYYY-MM-DD
    tipo:        str            # Ligação | WhatsApp | Email | Reunião | Outro
    resultado:   str
    proximo_fup: Optional[str] = None   # ISO date


def _contatos_col(doc_id: str):
    return _pipeline_col().document(doc_id).collection("contatos")


@app.get("/api/pipeline/{doc_id}/contatos")
async def pipeline_contatos_listar(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    """Lista o histórico de contatos de um cliente."""
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")

    try:
        ref = _pipeline_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")

        docs = _contatos_col(doc_id).order_by("data", direction="DESCENDING").stream()
        contatos = []
        for d in docs:
            c = d.to_dict() or {}
            c["id"] = d.id
            for k, v in c.items():
                if hasattr(v, "isoformat"):
                    c[k] = v.isoformat()
            contatos.append(c)
        return {"contatos": contatos}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"pipeline_contatos_listar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/pipeline/{doc_id}/contatos")
async def pipeline_contatos_adicionar(
    doc_id: str,
    body: ContatoLog,
    authorization: Optional[str] = Header(default=None),
):
    """Registra um contato no histórico e atualiza ultimo_contato/proximo_fup do cliente."""
    token_data    = await verificar_token(authorization)
    role          = token_data.get("role", "assessor")
    assessor_uid  = token_data.get("uid")
    assessor_name = token_data.get("assessor_name", "") or token_data.get("email", "")

    try:
        ref = _pipeline_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")

        entrada = body.dict()
        entrada["assessor_name"] = assessor_name
        entrada["created_at"]    = datetime.utcnow().isoformat()

        _contatos_col(doc_id).add(entrada)

        # Atualiza campos principais do cliente
        update_fields: dict = {
            "ultimo_contato": body.data,
            "updated_at":     datetime.utcnow().isoformat(),
        }
        if body.proximo_fup:
            update_fields["proximo_fup"] = body.proximo_fup

        ref.update(update_fields)
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"pipeline_contatos_adicionar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Pipeline: template Excel + importação em lote ────────────────────────────

PIPELINE_TEMPLATE_COLS = [
    "Nome", "Telefone", "Email", "Conta Antiga", "Origem",
    "AuC Global (R$)", "AuC BTG (R$)",
    "Ranking", "Pipe", "Status", "Observações",
]

RANKING_OPTS = "Com certeza | Possivelmente | Stand By | Desafio | Não Vem"
PIPE_OPTS    = "Quente | Morno | Frio | LEAD"


@app.get("/api/pipeline/template")
async def pipeline_template(
    authorization: Optional[str] = Header(default=None),
):
    """Retorna um arquivo .xlsx como template para carga em lote."""
    await verificar_token(authorization)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    header_fill = PatternFill("solid", fgColor="0C1A35")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    for col_idx, col_name in enumerate(PIPELINE_TEMPLATE_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3 linhas de exemplo
    exemplos = [
        ["João Silva", "11999990001", "joao@email.com", "", "Indicação",
         1500000, 0, "Com certeza", "Quente", "Primeiro contato feito", "Cliente VIP"],
        ["Maria Souza", "11999990002", "maria@email.com", "12345", "LinkedIn",
         500000, 200000, "Possivelmente", "Morno", "Aguardando retorno", ""],
        ["Pedro Lima",  "11999990003", "",               "",      "Evento",
         300000, None,   "Stand By",      "Frio",   "",                  "Amigo do cliente X"],
    ]
    example_fill = PatternFill("solid", fgColor="F0F9FF")
    for row_idx, row in enumerate(exemplos, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = example_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Linha de legenda (logo abaixo dos exemplos)
    legend_row = len(exemplos) + 2
    ws.cell(row=legend_row, column=1, value="⚠ Ranking válido:").font = Font(bold=True, color="92400E", size=10)
    ws.cell(row=legend_row, column=2, value=RANKING_OPTS).font = Font(color="92400E", size=10)
    ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=6)

    legend_row2 = legend_row + 1
    ws.cell(row=legend_row2, column=1, value="⚠ Pipe válido:").font = Font(bold=True, color="1E40AF", size=10)
    ws.cell(row=legend_row2, column=2, value=PIPE_OPTS).font = Font(color="1E40AF", size=10)
    ws.merge_cells(start_row=legend_row2, start_column=2, end_row=legend_row2, end_column=6)

    # Larguras
    widths = [28, 16, 28, 16, 20, 18, 18, 18, 12, 24, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pipeline_template.xlsx"},
    )


PIPELINE_COL_MAP = {
    "nome":        "Nome",
    "telefone":    "Telefone",
    "email":       "Email",
    "conta_antiga":"Conta Antiga",
    "origem":      "Origem",
    "auc_global":  "AuC Global (R$)",
    "auc_btg":     "AuC BTG (R$)",
    "ranking":     "Ranking",
    "pipe":        "Pipe",
    "status":      "Status",
    "observacoes": "Observações",
}

VALID_RANKING = {"Com certeza", "Possivelmente", "Stand By", "Desafio", "Não Vem"}
VALID_PIPE    = {"Quente", "Morno", "Frio", "LEAD"}


@app.post("/api/pipeline/import")
async def pipeline_importar(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    """Importa clientes em lote a partir de um arquivo .xlsx."""
    token_data    = await verificar_token(authorization)
    assessor_uid  = token_data.get("uid")
    assessor_name = token_data.get("assessor_name", "") or token_data.get("email", "")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler Excel: {e}")

    # Mapeamento inverso: coluna do Excel → campo interno
    inv = {v: k for k, v in PIPELINE_COL_MAP.items()}
    df.rename(columns=inv, inplace=True)

    erros   = []
    criados = 0
    col     = _pipeline_col()
    agora   = datetime.utcnow().isoformat()

    for i, row in df.iterrows():
        nome = str(row.get("nome", "") or "").strip()
        if not nome or nome.lower() == "nan":
            continue  # pula linhas vazias

        ranking = str(row.get("ranking", "") or "").strip()
        pipe    = str(row.get("pipe",    "") or "").strip()

        if ranking and ranking not in VALID_RANKING:
            erros.append(f"Linha {i+2}: Ranking inválido '{ranking}'")
            continue
        if pipe and pipe not in VALID_PIPE:
            erros.append(f"Linha {i+2}: Pipe inválido '{pipe}'")
            continue

        def _num(v):
            try:
                return float(str(v).replace(",", ".")) if v and str(v).strip() not in ("", "nan") else None
            except Exception:
                return None

        auc_global = _num(row.get("auc_global"))
        auc_btg    = _num(row.get("auc_btg"))

        dados = {
            "nome":          nome,
            "telefone":      str(row.get("telefone") or "").strip() or None,
            "email":         str(row.get("email")    or "").strip() or None,
            "conta_antiga":  str(row.get("conta_antiga") or "").strip() or None,
            "origem":        str(row.get("origem")   or "").strip() or None,
            "auc_global":    auc_global,
            "auc_btg":       auc_btg,
            "auc_potencial": round((auc_global or 0) - (auc_btg or 0), 2),
            "ranking":       ranking or None,
            "pipe":          pipe    or None,
            "status":        str(row.get("status")      or "").strip() or None,
            "observacoes":   str(row.get("observacoes") or "").strip() or None,
            "assessor_uid":  assessor_uid,
            "assessor_name": assessor_name,
            "created_at":    agora,
            "updated_at":    agora,
        }
        try:
            col.add(dados)
            criados += 1
        except Exception as e:
            erros.append(f"Linha {i+2}: {e}")

    return {"ok": True, "criados": criados, "erros": erros}


# ══════════════════════════════════════════════════════════════════════════════
# Produtos Manuais (admin) — Firestore + GCS
# ══════════════════════════════════════════════════════════════════════════════

def _prod_manuais_col():
    return get_fs().collection("produtos_manuais")


def _ser_prod_manual(doc) -> dict:
    d = doc.to_dict() or {}
    d["id"] = doc.id
    for k, v in d.items():
        if hasattr(v, "isoformat"):
            d[k] = v.isoformat()
    return d


@app.get("/api/produtos-manuais")
async def listar_produtos_manuais(
    authorization: Optional[str] = Header(default=None),
):
    """Lista todos os produtos incluídos manualmente."""
    await verificar_token(authorization)
    try:
        docs = _prod_manuais_col().stream()
        return {"produtos": [_ser_prod_manual(d) for d in docs]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/produtos-manuais")
async def criar_produto_manual(
    nome:            str            = Form(...),
    tipo:            str            = Form(...),
    emissor:         Optional[str]  = Form(default=None),
    rentabilidade:   Optional[str]  = Form(default=None),
    indexador:       Optional[str]  = Form(default=None),
    min_aplicacao:   Optional[str]  = Form(default=None),
    liquidez:        Optional[str]  = Form(default=None),
    prazo_meses:     Optional[str]  = Form(default=None),
    data_vencimento: Optional[str]  = Form(default=None),
    risco:           Optional[str]  = Form(default=None),
    descricao:       Optional[str]  = Form(default=None),
    link_externo:    Optional[str]  = Form(default=None),
    arquivo:         Optional[UploadFile] = File(default=None),
    authorization:   Optional[str]  = Header(default=None),
):
    """Cria produto manual — apenas admin."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso restrito a admins.")

    gcs_path    = None
    arquivo_nome = None
    arquivo_tipo = None
    if arquivo and arquivo.filename:
        content = await arquivo.read()
        safe    = re.sub(r"[^\w\-_\.]", "_", arquivo.filename)
        gcs_path    = f"{GCS_PREFIX_PRODUTOS}{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{safe}"
        arquivo_tipo = arquivo.content_type or "application/octet-stream"
        blob = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
        blob.upload_from_string(content, content_type=arquivo_tipo)
        arquivo_nome = arquivo.filename

    def _num(v):
        try:
            return float(str(v).replace(",", ".")) if v else None
        except Exception:
            return None

    dados = {
        "nome":            nome,
        "tipo":            tipo,
        "emissor":         emissor        or None,
        "rentabilidade":   rentabilidade  or None,
        "indexador":       indexador      or None,
        "min_aplicacao":   _num(min_aplicacao),
        "liquidez":        liquidez       or None,
        "prazo_meses":     int(prazo_meses) if prazo_meses and prazo_meses.strip().isdigit() else None,
        "data_vencimento": data_vencimento or None,
        "risco":           risco          or None,
        "descricao":       descricao      or None,
        "link_externo":    link_externo   or None,
        "arquivo_gcs":     gcs_path,
        "arquivo_nome":    arquivo_nome,
        "arquivo_tipo":    arquivo_tipo,
        "criado_por":      token_data.get("assessor_name") or token_data.get("email", ""),
        "created_at":      datetime.utcnow().isoformat(),
    }
    try:
        ref = _prod_manuais_col().add(dados)
        return {"ok": True, "id": ref[1].id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/produtos-manuais/{doc_id}")
async def deletar_produto_manual(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    """Remove produto manual — apenas admin."""
    token_data = await verificar_token(authorization)
    if token_data.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acesso restrito a admins.")

    ref = _prod_manuais_col().document(doc_id)
    doc = ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Produto não encontrado.")

    gcs_path = (doc.to_dict() or {}).get("arquivo_gcs")
    if gcs_path:
        try:
            gcs_client.bucket(GCS_BUCKET).blob(gcs_path).delete()
        except Exception:
            pass

    ref.delete()
    return {"ok": True}


@app.get("/api/produtos-manuais/{doc_id}/arquivo")
async def download_produto_manual_arquivo(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    """Retorna (stream) o arquivo anexado ao produto — todos os assessores podem baixar."""
    await verificar_token(authorization)

    ref = _prod_manuais_col().document(doc_id)
    doc = ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Produto não encontrado.")

    d        = doc.to_dict() or {}
    gcs_path = d.get("arquivo_gcs")
    fname    = d.get("arquivo_nome", "arquivo")
    ct       = d.get("arquivo_tipo", "application/octet-stream")

    if not gcs_path:
        raise HTTPException(status_code=404, detail="Nenhum arquivo anexado.")

    blob = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
    buf  = io.BytesIO()
    blob.download_to_file(buf)
    buf.seek(0)

    # PDFs abrem inline; outros fazem download
    disposition = "inline" if "pdf" in ct else f'attachment; filename="{fname}"'
    return StreamingResponse(buf, media_type=ct, headers={"Content-Disposition": disposition})


# ══════════════════════════════════════════════════════════════════════════════
# CHAMADOS (suporte / melhorias)
# ══════════════════════════════════════════════════════════════════════════════

def _chamados_col():
    return get_fs().collection("chamados")

VALID_TIPO_CHAMADO      = {"Melhoria", "Bug", "Dúvida"}
VALID_PRIORIDADE_CHAMADO = {"Baixa", "Média", "Alta"}
VALID_STATUS_CHAMADO    = {"Aberto", "Em análise", "Concluído"}


@app.post("/api/chamados")
async def abrir_chamado(
    titulo:      str       = Form(...),
    tipo:        str       = Form(...),
    prioridade:  str       = Form(...),
    descricao:   str       = Form(...),
    arquivo:     Optional[UploadFile] = File(default=None),
    authorization: Optional[str] = Header(default=None),
):
    """Abre um novo chamado. Qualquer usuário autenticado pode abrir."""
    payload = await verificar_token(authorization)
    uid            = payload.get("uid", "")
    assessor_name  = payload.get("assessor_name", uid)

    if tipo not in VALID_TIPO_CHAMADO:
        raise HTTPException(status_code=400, detail=f"Tipo inválido: {tipo}")
    if prioridade not in VALID_PRIORIDADE_CHAMADO:
        raise HTTPException(status_code=400, detail=f"Prioridade inválida: {prioridade}")

    arquivo_gcs  = None
    arquivo_nome = None
    arquivo_tipo = None

    if arquivo and arquivo.filename:
        content   = await arquivo.read()
        safe      = re.sub(r"[^\w.\-]", "_", arquivo.filename)
        gcs_path  = f"{GCS_PREFIX_CHAMADOS}{uid}/{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{safe}"
        blob      = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
        blob.upload_from_string(content, content_type=arquivo.content_type or "application/octet-stream")
        arquivo_gcs  = gcs_path
        arquivo_nome = arquivo.filename
        arquivo_tipo = arquivo.content_type or "application/octet-stream"

    dados = {
        "titulo":       titulo.strip(),
        "tipo":         tipo,
        "prioridade":   prioridade,
        "descricao":    descricao.strip(),
        "status":       "Aberto",
        "criado_por":   assessor_name,
        "uid":          uid,
        "created_at":   datetime.utcnow().isoformat(),
        "arquivo_gcs":  arquivo_gcs,
        "arquivo_nome": arquivo_nome,
        "arquivo_tipo": arquivo_tipo,
    }
    ref, doc = _chamados_col().add(dados)
    return {"id": doc.id, **dados}


@app.get("/api/chamados")
async def listar_chamados(
    authorization: Optional[str] = Header(default=None),
):
    """Admin vê todos os chamados. Assessor vê apenas os seus."""
    payload    = await verificar_token(authorization)
    uid        = payload.get("uid", "")
    role       = payload.get("role", "assessor")
    is_admin   = role == "admin"

    col = _chamados_col()
    if is_admin:
        docs = col.stream()
    else:
        docs = col.where("uid", "==", uid).stream()

    result = []
    for d in docs:
        item = d.to_dict() or {}
        item["id"] = d.id
        result.append(item)

    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return result


@app.patch("/api/chamados/{doc_id}")
async def atualizar_status_chamado(
    doc_id: str,
    body:   dict,
    authorization: Optional[str] = Header(default=None),
):
    """Atualiza status do chamado (somente admin)."""
    payload  = await verificar_token(authorization)
    uid      = payload.get("uid", "")
    role     = payload.get("role", "assessor")
    is_admin = role == "admin"
    if not is_admin:
        raise HTTPException(status_code=403, detail="Apenas admins podem alterar status.")

    novo_status = body.get("status")
    if novo_status not in VALID_STATUS_CHAMADO:
        raise HTTPException(status_code=400, detail=f"Status inválido: {novo_status}")

    ref = _chamados_col().document(doc_id)
    if not ref.get().exists:
        raise HTTPException(status_code=404, detail="Chamado não encontrado.")

    ref.update({"status": novo_status})
    return {"id": doc_id, "status": novo_status}


# ── Aprovações por Email ──────────────────────────────────────────────────────

class AprovacaoEmail(BaseModel):
    cliente_nome:  str
    cliente_email: str
    tipo:          str            # tesouro_direto | fundos_carencia
    data_inicio:   str            # YYYY-MM-DD
    data_fim:      str            # YYYY-MM-DD
    observacao:    Optional[str] = None

GCS_PREFIX_APROVACOES = "aprovacoes/"

def _aprovacoes_col():
    return fb_firestore.client().collection("aprovacoes")

def _ser_aprovacao(doc) -> dict:
    d = doc.to_dict() or {}
    d["id"] = doc.id
    return d


@app.get("/api/aprovacoes")
async def aprovacoes_listar(
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        col = _aprovacoes_col()
        if role == "admin":
            docs = col.order_by("data_fim").stream()
        else:
            docs = col.where("assessor_uid", "==", assessor_uid).order_by("data_fim").stream()
        return {"aprovacoes": [_ser_aprovacao(d) for d in docs]}
    except Exception as e:
        log.error(f"aprovacoes_listar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/aprovacoes")
async def aprovacoes_criar(
    body: AprovacaoEmail,
    authorization: Optional[str] = Header(default=None),
):
    token_data    = await verificar_token(authorization)
    assessor_uid  = token_data.get("uid")
    assessor_name = token_data.get("assessor_name", "") or token_data.get("email", "")
    dados = body.dict()
    dados["assessor_uid"]   = assessor_uid
    dados["assessor_name"]  = assessor_name
    dados["criado_em"]      = datetime.utcnow().isoformat()
    dados["resposta_gcs"]   = None
    dados["resposta_nome"]  = None
    dados["resposta_data"]  = None
    try:
        ref = _aprovacoes_col().add(dados)
        return {"ok": True, "id": ref[1].id}
    except Exception as e:
        log.error(f"aprovacoes_criar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/aprovacoes/{doc_id}")
async def aprovacoes_atualizar(
    doc_id: str,
    body: AprovacaoEmail,
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        ref = _aprovacoes_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Aprovação não encontrada.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")
        dados = {k: v for k, v in body.dict().items() if v is not None}
        dados["updated_at"] = datetime.utcnow().isoformat()
        ref.update(dados)
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"aprovacoes_atualizar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/aprovacoes/{doc_id}")
async def aprovacoes_deletar(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        ref = _aprovacoes_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Aprovação não encontrada.")
        d = doc.to_dict() or {}
        if role != "admin" and d.get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")
        # Remove arquivo GCS se existir
        if d.get("resposta_gcs"):
            try:
                gcs_client.bucket(GCS_BUCKET).blob(d["resposta_gcs"]).delete()
            except Exception:
                pass
        ref.delete()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"aprovacoes_deletar: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/aprovacoes/{doc_id}/resposta")
async def aprovacoes_upload_resposta(
    doc_id: str,
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        ref = _aprovacoes_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Aprovação não encontrada.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")

        # Remove arquivo anterior se existir
        old = doc.to_dict().get("resposta_gcs")
        if old:
            try:
                gcs_client.bucket(GCS_BUCKET).blob(old).delete()
            except Exception:
                pass

        fname    = file.filename or "resposta"
        gcs_path = f"{GCS_PREFIX_APROVACOES}{assessor_uid}/{doc_id}/{fname}"
        conteudo = await file.read()
        blob     = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
        blob.upload_from_string(conteudo, content_type=file.content_type or "application/octet-stream")

        ref.update({
            "resposta_gcs":  gcs_path,
            "resposta_nome": fname,
            "resposta_tipo": file.content_type or "application/octet-stream",
            "resposta_data": datetime.utcnow().strftime("%Y-%m-%d"),
        })
        return {"ok": True, "gcs_path": gcs_path}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"aprovacoes_upload_resposta: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/aprovacoes/{doc_id}/resposta")
async def aprovacoes_remover_resposta(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        ref = _aprovacoes_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Aprovação não encontrada.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")
        gcs_path = doc.to_dict().get("resposta_gcs")
        if gcs_path:
            try:
                gcs_client.bucket(GCS_BUCKET).blob(gcs_path).delete()
            except Exception:
                pass
        ref.update({"resposta_gcs": None, "resposta_nome": None, "resposta_data": None, "resposta_tipo": None})
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"aprovacoes_remover_resposta: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/aprovacoes/{doc_id}/resposta")
async def aprovacoes_download_resposta(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    token_data   = await verificar_token(authorization)
    role         = token_data.get("role", "assessor")
    assessor_uid = token_data.get("uid")
    try:
        ref = _aprovacoes_col().document(doc_id)
        doc = ref.get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="Aprovação não encontrada.")
        if role != "admin" and doc.to_dict().get("assessor_uid") != assessor_uid:
            raise HTTPException(status_code=403, detail="Sem permissão.")
        d        = doc.to_dict() or {}
        gcs_path = d.get("resposta_gcs")
        fname    = d.get("resposta_nome", "resposta")
        ct       = d.get("resposta_tipo", "application/octet-stream")
        if not gcs_path:
            raise HTTPException(status_code=404, detail="Nenhum arquivo de resposta.")
        blob = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
        buf  = io.BytesIO()
        blob.download_to_file(buf)
        buf.seek(0)
        disposition = "inline" if "pdf" in ct else f'attachment; filename="{fname}"'
        return StreamingResponse(buf, media_type=ct, headers={"Content-Disposition": disposition})
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"aprovacoes_download_resposta: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Notícias: refresh manual ──────────────────────────────────────────────────

def _buscar_noticia_api(categoria: dict, urls_vistos: set) -> dict | None:
    """Busca 1 notícia válida da categoria via NewsAPI."""
    if not NEWSAPI_KEY:
        return None
    try:
        resp = requests.get(
            NEWSAPI_URL_TUDO,
            params={"q": categoria["query"], "language": "pt",
                    "sortBy": "publishedAt", "pageSize": 5, "apiKey": NEWSAPI_KEY},
            timeout=10,
        )
    except Exception as e:
        log.warning(f"NewsAPI rede ({categoria['nome']}): {e}")
        return None
    if resp.status_code != 200:
        log.warning(f"NewsAPI {resp.status_code} ({categoria['nome']}): {resp.text[:200]}")
        return None
    for a in resp.json().get("articles", []):
        titulo    = (a.get("title")       or "").strip()
        descricao = (a.get("description") or "").strip()
        url       = (a.get("url")         or "").strip()
        if not titulo or titulo == "[Removed]": continue
        if not descricao or descricao == "[Removed]": continue
        if url in urls_vistos: continue
        urls_vistos.add(url)
        return {
            "categoria": categoria["nome"], "slug": categoria["slug"],
            "titulo": titulo, "descricao": descricao, "url": url,
            "fonte":  (a.get("source") or {}).get("name", ""),
            "data_pub": a.get("publishedAt", ""),
        }
    return None


@app.post("/api/noticias/refresh")
async def noticias_refresh(
    authorization: Optional[str] = Header(default=None),
):
    """Atualiza as notícias manualmente. Limite: NOTICIAS_LIMITE_MANUAL por dia."""
    await verificar_token(authorization)

    if not NEWSAPI_KEY:
        raise HTTPException(status_code=503, detail="NewsAPI não configurada neste serviço. Adicione NEWSAPI_KEY ao Cloud Run.")

    fs     = fb_firestore.client()
    hoje   = datetime.utcnow().strftime("%Y-%m-%d")

    # ── Controle de limite diário ─────────────────────────────────────────────
    ctrl_ref  = fs.collection("noticias").document("_controle")
    ctrl_snap = ctrl_ref.get()
    ctrl_data = ctrl_snap.to_dict() if ctrl_snap.exists else {}

    if ctrl_data.get("data") == hoje:
        count = int(ctrl_data.get("manual_count", 0))
    else:
        count = 0   # novo dia — zera

    if count >= NOTICIAS_LIMITE_MANUAL:
        raise HTTPException(
            status_code=429,
            detail=f"Limite diário de {NOTICIAS_LIMITE_MANUAL} atualizações manuais atingido. Tente amanhã ou aguarde o job automático."
        )

    # ── Determina período (horário de Brasília) ───────────────────────────────
    from datetime import timezone, timedelta as td
    agora_brt = datetime.now(timezone(td(hours=-3)))
    periodo   = "manha" if agora_brt.hour < 12 else "tarde"

    # ── Busca notícias ────────────────────────────────────────────────────────
    urls_vistos: set = set()
    itens: list = []

    for cat in NOTICIAS_CATEGORIAS:
        noticia = _buscar_noticia_api(cat, urls_vistos)
        if noticia:
            itens.append(noticia)
        time.sleep(1)   # respeita rate limit NewsAPI (1 req/s)

    if not itens:
        raise HTTPException(status_code=502, detail="Nenhuma notícia encontrada na NewsAPI.")

    # ── Salva no Firestore ────────────────────────────────────────────────────
    fs.collection("noticias").document(periodo).set({
        "periodo":   periodo,
        "gerado_em": datetime.utcnow().isoformat(),
        "itens":     itens,
    })

    # ── Atualiza contador ─────────────────────────────────────────────────────
    novo_count = count + 1
    ctrl_ref.set({"data": hoje, "manual_count": novo_count})

    return {
        "ok":        True,
        "itens":     itens,
        "usos_hoje": novo_count,
        "limite":    NOTICIAS_LIMITE_MANUAL,
    }


@app.get("/api/noticias/controle")
async def noticias_controle(
    authorization: Optional[str] = Header(default=None),
):
    """Retorna o uso manual de hoje para exibir no dashboard."""
    await verificar_token(authorization)
    fs    = fb_firestore.client()
    hoje  = datetime.utcnow().strftime("%Y-%m-%d")
    snap  = fs.collection("noticias").document("_controle").get()
    data  = snap.to_dict() if snap.exists else {}
    count = int(data.get("manual_count", 0)) if data.get("data") == hoje else 0
    return {"usos_hoje": count, "limite": NOTICIAS_LIMITE_MANUAL}


@app.get("/api/chamados/{doc_id}/arquivo")
async def download_chamado_arquivo(
    doc_id: str,
    authorization: Optional[str] = Header(default=None),
):
    """Retorna o arquivo anexado ao chamado."""
    payload  = await verificar_token(authorization)
    uid      = payload.get("uid", "")
    role     = payload.get("role", "assessor")
    is_admin = role == "admin"

    ref = _chamados_col().document(doc_id)
    doc = ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Chamado não encontrado.")

    d = doc.to_dict() or {}

    # Assessor só pode baixar anexo do próprio chamado
    if not is_admin and d.get("uid") != uid:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    gcs_path = d.get("arquivo_gcs")
    fname    = d.get("arquivo_nome", "arquivo")
    ct       = d.get("arquivo_tipo", "application/octet-stream")

    if not gcs_path:
        raise HTTPException(status_code=404, detail="Nenhum arquivo anexado.")

    blob = gcs_client.bucket(GCS_BUCKET).blob(gcs_path)
    buf  = io.BytesIO()
    blob.download_to_file(buf)
    buf.seek(0)

    disposition = "inline" if "pdf" in ct else f'attachment; filename="{fname}"'
    return StreamingResponse(buf, media_type=ct, headers={"Content-Disposition": disposition})


# ── Movimentações BTG ─────────────────────────────────────────────────────────

@app.post("/api/movimentacoes/solicitar/{conta}")
async def movimentacoes_solicitar(
    conta: str,
    authorization: Optional[str] = Header(default=None),
):
    """
    Solicita histórico de movimentações ao BTG via OAuth2.
    A resposta é assíncrona — o BTG chama o webhook-btg que salva no Firestore.
    """
    token_data = await verificar_token(authorization)
    uid        = token_data.get("uid", "")

    # Obtém token BTG
    try:
        btg_token = _get_btg_token()
    except Exception as e:
        log.error(f"Erro ao obter token BTG: {e}")
        raise HTTPException(status_code=502, detail="Erro de autenticação com a BTG. Verifique BTG_CLIENT_ID/SECRET.")

    # Chama API BTG (resposta assíncrona — dados chegam via webhook)
    try:
        resp = requests.get(
            f"{BTG_MOV_URL}/{conta}",
            headers={"Authorization": f"Bearer {btg_token}"},
            timeout=30,
        )
        resp.raise_for_status()
    except requests.HTTPError as e:
        log.error(f"Erro BTG API movimentações conta {conta}: {e} — {getattr(e.response,'text','')}")
        raise HTTPException(status_code=502, detail=f"Erro ao solicitar dados BTG: {resp.status_code}")
    except Exception as e:
        log.error(f"Erro de comunicação BTG: {e}")
        raise HTTPException(status_code=502, detail="Erro de comunicação com a BTG.")

    # Registra solicitação no Firestore (status = aguardando)
    fs = fb_firestore.client()
    fs.collection("movimentacoes").document(conta).set({
        "conta":            conta,
        "uid_solicitante":  uid,
        "solicitado_em":    datetime.utcnow().isoformat() + "Z",
        "status":           "aguardando",
        "dados":            None,
        "atualizado_em":    None,
    }, merge=True)

    log.info(f"Movimentações solicitadas: conta={conta} uid={uid}")
    return {"ok": True, "conta": conta, "status": "aguardando",
            "msg": "Solicitação enviada. Os dados chegam em instantes via webhook."}


@app.get("/api/movimentacoes/{conta}")
async def movimentacoes_get(
    conta: str,
    authorization: Optional[str] = Header(default=None),
):
    """Retorna dados de movimentação salvos no Firestore para a conta."""
    await verificar_token(authorization)

    fs  = fb_firestore.client()
    doc = fs.collection("movimentacoes").document(conta).get()
    if not doc.exists:
        return {"conta": conta, "status": "sem_dados", "dados": None,
                "solicitado_em": None, "atualizado_em": None}

    d = doc.to_dict() or {}
    return {
        "conta":          conta,
        "status":         d.get("status", "sem_dados"),
        "dados":          d.get("dados"),
        "solicitado_em":  d.get("solicitado_em"),
        "atualizado_em":  d.get("atualizado_em"),
    }
